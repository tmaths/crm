from collections import defaultdict
from django.shortcuts import render
from django.contrib import messages
from django.http import HttpResponse, FileResponse, JsonResponse
from django.db.models import Q, Count, Case, When, IntegerField, Sum, Min
from django.conf import settings
from django.utils import timezone
from django.shortcuts import render, redirect
from django import forms
from .models import Usagereport, ReportGeneration, ENQUIRY_RATES, SubscriberProductRate, Subscriber, KeySubscriber
from datetime import date, timedelta, datetime
import calendar
import io
import os
import re
import zipfile
from copy import copy, deepcopy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import os.path
import uuid
import tempfile
import gc
from . import excel_builder
from django.urls import reverse
from django.contrib.auth.decorators import login_required,user_passes_test
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import logging
from django.core.cache import cache
import hashlib
from django.db.models import DateField
from django.db.models.functions import Cast
from django.db.models import Exists, OuterRef
import csv
from .pdf_generator import PDFReportGenerator

# Get an instance of a logger
logger = logging.getLogger(__name__)


def get_cached_subscriber_rates(subscriber_name, logger=None):
    if logger is None:
        logger = logging.getLogger(__name__)
    
    cache_key = f'rates_{subscriber_name.replace(" ", "_")}'
    cached_rates = cache.get(cache_key)
    
    if cached_rates is not None:
        logger.debug(f"Using cached rates for {subscriber_name}")
        return cached_rates
    
    # Fetch all rates for this subscriber from database
    logger.debug(f"Fetching rates from database for {subscriber_name}")
    custom_rates = SubscriberProductRate.objects.filter(
        subscriber_name__iexact=subscriber_name
    ).values('product_name', 'rate')
    
    # Build rates dictionary with defaults
    rates_dict = {}
    for rate_record in custom_rates:
        product_name = rate_record['product_name']
        rates_dict[product_name.lower()] = Decimal(str(rate_record['rate']))
    
    # Cache for 1 hour (3600 seconds)
    cache.set(cache_key, rates_dict, 3600)
    logger.debug(f"Cached {len(rates_dict)} rates for {subscriber_name}")
    
    return rates_dict


def get_rate_with_cache(subscriber_name, product_name, default_rate, cached_rates=None, logger=None):
    if logger is None:
        logger = logging.getLogger(__name__)
    
    # Get cached rates if not provided
    if cached_rates is None:
        cached_rates = get_cached_subscriber_rates(subscriber_name, logger)
    
    # Look up rate (case-insensitive)
    rate = cached_rates.get(product_name.lower())
    
    if rate is not None:
        return rate
    else:
        # Return default rate
        if not isinstance(default_rate, Decimal):
            return Decimal(str(default_rate))
        return default_rate

# Old template cache functions removed - now using JSON-based builder

def write_excel_to_temp_file(workbook, filename):
    """Write Excel workbook to a temporary file and return the file path."""
    try:
        # Create a temporary file
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx', prefix=f'{filename}_')
        os.close(temp_fd)  # Close the file descriptor
        
        # Save the workbook to the temporary file
        workbook.save(temp_path)
        
        # Clear the workbook from memory
        workbook.close()
        del workbook
        gc.collect()
        
        return temp_path
    except Exception as e:
        logger.error(f"Error writing Excel file to temp: {str(e)}")
        raise

def create_merged_cell_map(worksheet):
    """Create a lookup map for merged cells to optimize cell operations."""
    merged_map = {}
    for merged_range in worksheet.merged_cells.ranges:
        range_str = str(merged_range)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(row, col)] = {
                    'range': range_str,
                    'is_top_left': (row == merged_range.min_row and col == merged_range.min_col),
                    'min_row': merged_range.min_row,
                    'min_col': merged_range.min_col
                }
    return merged_map

def optimized_cell_assignment(ws, row, col, value, merged_map=None):
    """Optimized version of safe_cell_assignment using pre-calculated merged cell map."""
    if merged_map and (row, col) in merged_map:
        merge_info = merged_map[(row, col)]
        if merge_info['is_top_left']:
            ws.cell(row=row, column=col).value = value
        else:
            ws.cell(row=merge_info['min_row'], column=merge_info['min_col']).value = value
    else:
        # Not a merged cell, write directly
        ws.cell(row=row, column=col).value = value


def get_key_subscribers_list():
    """Fetch key subscribers from database instead of hardcoded list."""
    try:
        return list(KeySubscriber.objects.values_list('subscriber_name', flat=True))
    except Exception as e:
        # If table doesn't exist (migration not run), return empty list
        logger.warning(f"Could not fetch key subscribers from database: {e}")
        return []


def safe_cell_assignment(ws, row, col, value):
    """Helper function to safely assign values to cells, handling merged cells."""
    write_to_cell(ws, row, col, value)

def write_to_cell(ws, row, col, value):
    """
    Safely writes a value to a cell, handling merged cells and preserving formatting.
    
    When using the direct cell assignment approach like in VBA, we need to 
    handle merged cells specially as they are read-only in openpyxl.
    """
    coordinate = ws.cell(row=row, column=col).coordinate
    is_merged = False
    merged_range_to_restore = None
    
    # Store the original cell style before unmerging
    original_cell = ws.cell(row=row, column=col)
    original_style = original_cell._style
    original_number_format = original_cell.number_format
    
    # Special handling for product name header (row 32)
    if row == 32 and col == 4:  # D32
        # Try additional columns for row 32 as it might be a merged cell
        for try_col in range(1, 10):  # Try columns A through I
            try:
                ws.cell(row=row, column=try_col).value = value
            except:
                pass
    
    # Check if the cell is in a merged range
    for merged_range in list(ws.merged_cells.ranges):
        if coordinate in merged_range:
            is_merged = True
            merged_range_to_restore = str(merged_range)
            ws.unmerge_cells(merged_range_to_restore)
            break
    
    # Now we can safely set the value
    target_cell = ws.cell(row=row, column=col)
    target_cell.value = value
    
  
    target_cell._style = original_style
    if original_number_format != 'General':
        target_cell.number_format = original_number_format 
    if is_merged and merged_range_to_restore:
        ws.merge_cells(merged_range_to_restore)


def safe_cell_assignment(ws, row, col, value):
    """Helper function to safely assign values to cells, handling merged cells."""
    write_to_cell(ws, row, col, value)

# OPTIMIZED: Batch writing function for better performance
def batch_write_row(ws, row, col_values, preserve_merges=True):
    if not col_values:
        return
    # Store merged ranges that need to be preserved
    merges_to_restore = []
    if preserve_merges:
        for col in col_values.keys():
            coordinate = ws.cell(row=row, column=col).coordinate
            for merged_range in list(ws.merged_cells.ranges):
                if coordinate in merged_range:
                    merge_str = str(merged_range)
                    if merge_str not in merges_to_restore:
                        merges_to_restore.append(merge_str)
                        ws.unmerge_cells(merge_str)
                    break
    

    for col, value in col_values.items():
        cell = ws.cell(row=row, column=col)
        cell.value = value
    
    for merge_str in merges_to_restore:
        ws.merge_cells(merge_str)

def batch_write_rows(ws, start_row, rows_data, template_row=None):
    """
    Write multiple rows with proper styling from template.
    Now uses excel_builder for consistent formatting.
    """
    for i, row_data in enumerate(rows_data):
        current_row = start_row + i
        
        # Use excel_builder to clone template row with proper styling and merges
        if template_row:
            excel_builder.clone_template_row_to_data_row(
                ws, 
                template_row, 
                current_row, 
                value_overrides=row_data
            )
        else:
            # Fallback to simple write if no template
            for col, value in row_data.items():
                ws.cell(row=current_row, column=col).value = value

@login_required
def home(request):
    today = date.today()
    first_day_of_month = today.replace(day=1)
    
    # Calculate first day of next month
    if today.month == 12:
        first_day_next_month = date(today.year + 1, 1, 1)
    else:
        first_day_next_month = date(today.year, today.month + 1, 1)
    
    # Use Django ORM to fetch distinct subscriber names
    start_date = first_day_of_month
    end_date = first_day_next_month
    
    subscribers = Usagereport.objects.filter(
        DetailsViewedDate__gte=start_date,
        DetailsViewedDate__lt=end_date
    ).values_list('SubscriberName', flat=True).distinct().order_by('SubscriberName')

    # Format dates as strings for the template
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    context = {
        'subscribers': subscribers,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }
    return render(request, 'bulkrep/home.html', context)

def clean_filename(filename):
    """Clean subscriber name for a valid filename, similar to VBA script."""
    # Replace invalid characters with hyphens
    invalid_chars = r'[\/\\\:\*\?"<>\|]'
    return re.sub(invalid_chars, '-', filename)

@login_required
def single_report(request):
    """View for generating a single report."""
    # PERFORMANCE: Track execution time
    import time
    start_time = time.time()
    
    # --- Initial Setup (Largely Unchanged) ---
    today = date.today()
    first_day_of_month = today.replace(day=1)

    if today.month == 12:
        first_day_next_month = date(today.year + 1, 1, 1)
    else:
        first_day_next_month = date(today.year, today.month + 1, 1)

    report_gen = None
    subscribers = Usagereport.objects.values_list('SubscriberName', flat=True).distinct().order_by('SubscriberName')

    context = {
        'subscribers': subscribers,
        'start_date': first_day_of_month.strftime('%Y-%m-%d'),
        'end_date': first_day_next_month.strftime('%Y-%m-%d'),
    }

    if request.method == 'POST':
        print(f"\n{'='*60}")
        print(f"SINGLE REPORT GENERATION STARTED")
        print(f"{'='*60}")
        
        # --- Form Data & Report Tracking (Unchanged) ---
        subscriber_id = request.POST.get('subscriber_id')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        include_bills = request.POST.get('include_bills') == 'on'
        include_products = request.POST.get('include_products') == 'on'
        output_format = request.POST.get('output_format', 'excel')
        
        print(f"Subscriber: {subscriber_id}")
        print(f"Date Range: {start_date_str} to {end_date_str}")
        print(f"Include Bills: {include_bills}")
        print(f"Include Products: {include_products}")
        print(f"Output Format: {output_format}")

        if subscriber_id:
            report_gen = ReportGeneration.objects.create(
                user=request.user, generator=request.user.username, report_type='single',
                status='in_progress', subscriber_name=subscriber_id,
                from_date=start_date_str, to_date=end_date_str
            )

        if not subscriber_id:
            messages.error(request, "Please select a subscriber.")
            return render(request, 'bulkrep/single_report.html', context)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            start_date_display = start_date.strftime('%d/%m/%Y')
            end_date_display = end_date.strftime('%d/%m/%Y')
            print(f"Date parsing completed")
        except (ValueError, TypeError) as e:
            messages.error(request, f"Invalid date format: {str(e)}")
            return render(request, 'bulkrep/single_report.html', context)

        # --- OPTIMIZED: Fetch all product counts in a single query ---
        summary_bills = {}
        if include_bills:
            print(f"\n[1/5] Fetching billing data...")
            bill_start = time.time()
            queryset = Usagereport.objects.filter(
                DetailsViewedDate__gte=start_date,
                DetailsViewedDate__lte=end_date,
                SubscriberName=subscriber_id
            )
            summary_bills = queryset.aggregate(
                consumer_snap_check=Count(Case(When(ProductName__icontains='Snap Check', then=1))),
                consumer_basic_trace=Count(Case(When(ProductName__icontains='Basic Trace', then=1))),
                consumer_basic_credit=Count(Case(When(ProductName__icontains='Basic Credit', then=1))),
                consumer_detailed_credit=Count(Case(When(Q(ProductName__icontains='Consumer Detailed Credit') & ~Q(ProductName__icontains='X-SCore'), then=1))),
                xscore_consumer_detailed_credit=Count(Case(When(ProductName__icontains='X-SCore Consumer Detailed Credit', then=1))),
                commercial_basic_trace=Count(Case(When(ProductName__icontains='Commercial Basic Trace', then=1))),
                commercial_detailed_credit=Count(Case(When(ProductName__icontains='Commercial detailed Credit', then=1))),
                enquiry_report=Count(Case(When(ProductName__icontains='Enquiry Report', then=1))),
                consumer_dud_cheque=Count(Case(When(ProductName__icontains='Consumer Dud Cheque', then=1))),
                commercial_dud_cheque=Count(Case(When(ProductName__icontains='Commercial Dud Cheque', then=1))),
                director_basic_report=Count(Case(When(ProductName__icontains='Director Basic Report', then=1))),
                director_detailed_report=Count(Case(When(ProductName__icontains='Director Detailed Report', then=1))),
            )
            print(f"   Billing data fetched in {time.time() - bill_start:.2f}s")

        # Query for product details using Django ORM
        if include_products:
            print(f"\n[2/5] Fetching product data...")
            product_start = time.time()
            product_data = Usagereport.objects.filter(
                DetailsViewedDate__gte=start_date,
                DetailsViewedDate__lte=end_date,
                SubscriberName=subscriber_id
            ).order_by('ProductName', 'DetailsViewedDate').values(
                'SubscriberName', 'SystemUser', 'SearchIdentity', 'SubscriberEnquiryDate',
                'SearchOutput', 'DetailsViewedDate', 'ProductInputed', 'ProductName'
            )
            
            # Group by ProductName
            product_sections = {}
            for record in product_data:
                product_name = record['ProductName']
                if product_name not in product_sections:
                    product_sections[product_name] = []
                product_sections[product_name].append(record)
            
            product_data = list(product_data)
            print(f"   Found {len(product_data)} product records in {time.time() - product_start:.2f}s")
        else:
            product_sections = {}
            product_data = []

        if not product_data and include_products:
            messages.warning(request, f"No data found for subscriber {subscriber_id} between {start_date_display} and {end_date_display}.")
            return render(request, 'bulkrep/single_report.html', context)

        # === PDF GENERATION PATH ===
        if output_format == 'pdf':
            try:
                print(f"\n[3/4] Generating PDF report...")
                pdf_start = time.time()
                
                # Prepare summary_bills data with rates for PDF
                pdf_summary_bills = {}
                if include_bills:
                    # Get all custom rates for the subscriber at once
                    custom_rates_qs = SubscriberProductRate.objects.filter(subscriber_name__iexact=subscriber_id)
                    custom_rates_lookup = {rate.product_name.lower(): rate.rate for rate in custom_rates_qs}

                    def get_rate(product_name_key, product_name_display):
                        default_rate = ENQUIRY_RATES.get(product_name_key, Decimal('0.00'))
                        retrieved_rate = custom_rates_lookup.get(product_name_display.lower(), default_rate)
                        try:
                            return Decimal(retrieved_rate)
                        except (ValueError, TypeError, InvalidOperation):
                            return default_rate
                    
                    # Add rates to summary_bills for PDF
                    for key, name in [
                        ('consumer_snap_check', 'Consumer Snap Check'),
                        ('consumer_basic_trace', 'Consumer Basic Trace'),
                        ('consumer_basic_credit', 'Consumer Basic Credit'),
                        ('consumer_detailed_credit', 'Consumer Detailed Credit'),
                        ('xscore_consumer_detailed_credit', 'X-Score Consumer Detailed Credit'),
                        ('commercial_basic_trace', 'Commercial Basic Trace'),
                        ('commercial_detailed_credit', 'Commercial Detailed Credit'),
                        ('enquiry_report', 'Enquiry Report'),
                        ('consumer_dud_cheque', 'Consumer Dud Cheque'),
                        ('commercial_dud_cheque', 'Commercial Dud Cheque'),
                        ('director_basic_report', 'Director Basic Report'),
                        ('director_detailed_report', 'Director Detailed Report'),
                    ]:
                        pdf_summary_bills[key] = summary_bills.get(key, 0)
                        pdf_summary_bills[f'{key}_rate'] = get_rate(key, name)
                
                # Generate PDF
                pdf_generator = PDFReportGenerator()
                month_year = start_date.strftime('%B%Y')
                clean_subscriber = clean_filename(subscriber_id)
                filename = f"{clean_subscriber}_{month_year}_{uuid.uuid4().hex[:8]}.pdf"
                single_reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'single')
                os.makedirs(single_reports_dir, exist_ok=True)
                file_path = os.path.join(single_reports_dir, filename)
                
                pdf_generator.generate_report(
                    subscriber_name=subscriber_id,
                    start_date_display=start_date_display,
                    end_date_display=end_date_display,
                    summary_bills=pdf_summary_bills if include_bills else None,
                    product_data=product_data if include_products else None,
                    username=request.user.username,
                    output_path=file_path
                )
                
                pdf_elapsed = time.time() - pdf_start
                total_elapsed = time.time() - start_time
                print(f"   ✓ PDF generated in {pdf_elapsed:.2f}s")
                print(f"\n{'='*60}")
                print(f"✓ PDF REPORT GENERATION COMPLETE")
                print(f"   Total Time: {total_elapsed:.2f}s")
                if include_bills:
                    print(f"   Billing Products: {sum(summary_bills.values())}")
                if include_products:
                    print(f"   Product Records: {len(product_data)}")
                print(f"{'='*60}\n")
                
                download_url = settings.MEDIA_URL + f'reports/single/{filename}'
                
                # Update report generation status to success
                if 'report_gen' in locals() and report_gen:
                    report_gen.status = 'success'
                    report_gen.completed_at = timezone.now()
                    report_gen.save()
                
                # PERFORMANCE: Log execution time
                execution_time = time.time() - start_time
                logger.info(f"PDF report generated in {execution_time:.2f} seconds for {subscriber_id}")
                    
                return render(request, 'bulkrep/download_ready.html', {
                    'download_url': download_url
                })
                
            except Exception as e:
                error_msg = f"Error generating PDF report: {str(e)}"
                logger.error(error_msg)
                messages.error(request, error_msg)
                if 'report_gen' in locals() and report_gen:
                    report_gen.status = 'failed'
                    report_gen.error_message = error_msg[:500]
                    report_gen.completed_at = timezone.now()
                    report_gen.save()
                return render(request, 'bulkrep/single_report.html', context)

        # === EXCEL GENERATION PATH ===
        # Generate Excel report using JSON-based builder
        try:
            print(f"\n[3/5] Creating workbook from JSON template...")
            wb_start = time.time()
            
            # Prepare dynamic values for the template
            value_overrides = excel_builder.prepare_dynamic_values(
                subscriber_id, start_date_display, end_date_display
            )
            
            # Create workbook from JSON structure (NOT from .xlsx template)
            wb = excel_builder.create_workbook_from_json(value_overrides)
            ws = wb.active
            print(f"   Workbook created in {time.time() - wb_start:.2f}s")
        except Exception as e:
            logger.error(f"Error creating workbook from JSON template: {str(e)}")
            messages.error(request, f"Error creating workbook from JSON template: {str(e)}")
            return render(request, 'bulkrep/single_report.html', context)
        
        print(f"\n[4/5] Writing data to Excel...")
        write_start = time.time()
        try:
            # Look for "Productname" cell to identify where to put the dynamic product name
            product_name_cell = None
            for row in range(30, 35):  # Search rows 30-34
                for col in range(1, 10):  # Search columns A-I
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and "product" in str(cell_value).lower():
                        try:
                            product_name_cell = (row, col)
                        except Exception as e:
                            logger.error(f"Error finding product name cell: {str(e)}")
                            messages.error(request, f"Error finding product name cell: {str(e)}")
                            return render(request, 'bulkrep/single_report.html', context)
                        break
                if product_name_cell:
                    break
            
            row2_merged_range = None
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == 2 and merged_range.max_row == 2:
                    row2_merged_range = merged_range
                    break
            
            if row2_merged_range:
                merged_range_str = str(row2_merged_range)
                ws.unmerge_cells(merged_range_str)
                original_cell = ws.cell(row=2, column=row2_merged_range.min_col)
                new_content = f"FirstCentral NIGERIA - BILLING DETAILS - {subscriber_id}"
                original_cell.value = new_content
                
                ws.merge_cells(merged_range_str)
            else:
                # Fallback to the original method if no merged range is found
                safe_cell_assignment(ws, 2, 5, subscriber_id)  # E2
            
            safe_cell_assignment(ws, 5, 4, f"BILLING DETAILS - {subscriber_id}")  # D5
            
            # For row 6, which is merged from B to Q, we need to set the value to the first cell in the merged range
            # First, find the merged range that contains row 6
            row6_merged_range = None
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == 6 and merged_range.max_row == 6:
                    row6_merged_range = merged_range
                    break
            
            # If we found a merged range for row 6, unmerge it, set the value, and remerge it
            date_range_text = f"REPORT GENERATED FOR RECORDS BETWEEN {start_date_display} and {end_date_display}"
            if row6_merged_range:
                merged_range_str = str(row6_merged_range)
                ws.unmerge_cells(merged_range_str)
                # Set the value to the first cell in the merged range (column B = 2)
                ws.cell(row=6, column=2).value = date_range_text
                # Remerge the cells
                ws.merge_cells(merged_range_str)
            else:
                # Fallback to the original method if no merged range is found
                safe_cell_assignment(ws, 6, 4, date_range_text)  # D6
                
               
            
            # --- OPTIMIZED: Billing Logic ---
            if include_bills:
                # Fetch rates for the subscriber and add to summary_bills
                cached_rates = get_cached_subscriber_rates(subscriber_id, logger)
                
                # Add rates to summary_bills dict
                for key, name in [
                    ('consumer_snap_check', 'Consumer Snap Check'),
                    ('consumer_basic_trace', 'Consumer Basic Trace'),
                    ('consumer_basic_credit', 'Consumer Basic Credit'),
                    ('consumer_detailed_credit', 'Consumer Detailed Credit'),
                    ('xscore_consumer_detailed_credit', 'X-Score Consumer Detailed Credit'),
                    ('commercial_basic_trace', 'Commercial Basic Trace'),
                    ('commercial_detailed_credit', 'Commercial Detailed Credit'),
                    ('enquiry_report', 'Enquiry Report'),
                    ('consumer_dud_cheque', 'Consumer Dud Cheque'),
                    ('commercial_dud_cheque', 'Commercial Dud Cheque'),
                    ('director_basic_report', 'Director Basic Report'),
                    ('director_detailed_report', 'Director Detailed Report'),
                ]:
                    default_rate = ENQUIRY_RATES.get(key, Decimal('0.00'))
                    summary_bills[f'{key}_rate'] = get_rate_with_cache(
                        subscriber_id, name, default_rate, cached_rates, logger
                    )
                
                # Use excel_builder to write billing data
                excel_builder.write_billing_data(ws, summary_bills, subscriber_id, logger)
            
            if include_products:
                start_row_offset = 36  # Initial start row for product sections on Sheet1
                current_sheet = ws
                sheet2 = wb["Sheet2"] if "Sheet2" in wb.sheetnames else None
                
                # Sort product_sections by product_name
                sorted_product_names = sorted(product_sections.keys())
                
                # First, find the original product name cell in the template (rows 32-35)
                product_name_cell = None
                for row in range(32, 36):  # Rows 32-35 (inclusive)
                    for col in range(1, 16):  # Assuming columns A-O are important
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and "product" in str(cell_value).lower():
                            product_name_cell = (row, col)
                            break
                    if product_name_cell:
                        break
                        
                # Save the template header structure (rows 32-35) for subsequent products
                header_template = []
                header_rows = (32, 35)  # Range of rows to copy for the header template
                for row in range(header_rows[0], header_rows[1] + 1):
                    row_data = []
                    for col in range(1, 16):  # Assuming columns A-O are important
                        cell = ws.cell(row=row, column=col)
                        # Store cell value and position only - we'll copy styles directly later
                        cell_info = {
                            'value': cell.value,
                            'position': (row, col)
                        }
                        # Check if cell is part of merged range
                        for m_range in ws.merged_cells.ranges:
                            if (row, col) == (m_range.min_row, m_range.min_col):
                                cell_info['merged'] = (m_range.max_row - m_range.min_row + 1, 
                                                      m_range.max_col - m_range.min_col + 1)
                                break
                        row_data.append(cell_info)
                    header_template.append(row_data)
                
                # Set data row start - will be used for first product                
                data_start_row = 36  # Initial start for data rows after template header
                lastProduct = ""
                
                # For each product, use appropriate header section
                for product_idx, product_name in enumerate(sorted_product_names):
                    product_records = product_sections[product_name]
                    product_start_time = time.time()
                    print(f"   → Writing product '{product_name}' ({len(product_records)} records)...")
                    
                    if product_idx == 0:
                        # For first product, use the existing template header
                        if product_name_cell:
                            row, col = product_name_cell
                            # Replace "Product Name" with actual first product name in template
                            safe_cell_assignment(ws, row, col, product_name)
                        current_sheet = ws
                        current_row_offset = data_start_row  # Start data at row 36
                        serial_number_base = data_start_row - 1
                    else:
                        # Add space between different products (extra spacing)
                        current_row_offset += 4  # Add significant spacing between products
                            
                        # Create new header for subsequent products
                        header_start_row = current_row_offset
                        header_clone_start = time.time()
                        print(f"      Cloning header for product {product_idx + 1}...")
                        
                        # Clone the header section for this product
                        for template_row_idx, template_row in enumerate(header_template):
                            target_row = header_start_row + template_row_idx
                        
                            # Ensure we're not past row limits
                            if current_sheet == ws and target_row > 1000000 and sheet2:
                                current_sheet = sheet2
                                # Reset for Sheet2
                                header_start_row = 13
                                target_row = header_start_row + template_row_idx
                        
                            # Unmerge any existing merged cells in the target area
                            for m_range in list(current_sheet.merged_cells.ranges):
                                if m_range.min_row <= target_row <= m_range.max_row:
                                    current_sheet.unmerge_cells(
                                        start_row=m_range.min_row, 
                                        start_column=m_range.min_col,
                                        end_row=m_range.max_row, 
                                        end_column=m_range.max_col
                                    )
                        
                            # Copy each cell from the template to the target area
                            for col_idx, cell_info in enumerate(template_row):
                                target_col = col_idx + 1
                                target_cell = current_sheet.cell(row=target_row, column=target_col)
                                
                                # Copy styles directly from the original cell
                                original_row, original_col = cell_info['position']
                                original_cell = ws.cell(row=original_row, column=original_col)
                                
                                # Copy cell format using openpyxl's built-in method
                                target_cell._style = copy(original_cell._style)
                                
                                # Set value using safe_cell_assignment to handle merged cells properly
                                if template_row_idx == product_name_cell[0] - header_rows[0] and \
                                   col_idx == product_name_cell[1] - 1:
                                    # For the product name cell, use the safe assignment method
                                    safe_cell_assignment(current_sheet, target_row, target_col, product_name)
                                elif cell_info['value'] is not None:
                                    # For other cells with values, use the safe assignment method
                                    safe_cell_assignment(current_sheet, target_row, target_col, cell_info['value'])
                                
                                # Recreate merged cells
                                if 'merged' in cell_info:
                                    rows, cols = cell_info['merged']
                                    current_sheet.merge_cells(
                                        start_row=target_row, 
                                        start_column=target_col,
                                        end_row=target_row + rows - 1, 
                                        end_column=target_col + cols - 1
                                    )
                        
                        # Update data start row to be after this new header
                        current_row_offset = header_start_row + (header_rows[1] - header_rows[0] + 1)
                        serial_number_base = current_row_offset - 1
                        
                        # Add header for the data section
                        safe_cell_assignment(current_sheet, current_row_offset - 1, 4, "Unique Tracking Number")
                        
                        header_clone_elapsed = time.time() - header_clone_start
                        print(f"      Header cloned in {header_clone_elapsed:.2f}s")

                    # Process data records for this product - OPTIMIZED
                    data_write_start = time.time()
                    
                    # OPTIMIZATION: Build all row data first, then batch write
                    all_rows_data = []
                    for record_idx, record in enumerate(product_records):
                        current_row = current_row_offset + record_idx
                        
                        # Switch to Sheet2 when reaching max row (like VBA)
                        if current_row > 1000000 and sheet2 and current_sheet != sheet2:
                            current_sheet = sheet2
                            current_row_offset = 13  # Reset row offset for Sheet2 as per VBA
                            serial_number_base = 12  # Reset serial number base for Sheet2
                            current_row = current_row_offset + record_idx  # Recalculate current_row for Sheet2
                        
                        # Format dates
                        subscriber_enquiry_date_str = ""
                        if record['SubscriberEnquiryDate']:
                            if isinstance(record['SubscriberEnquiryDate'], str):
                                subscriber_enquiry_date_str = record['SubscriberEnquiryDate']
                            else:
                                date_only = record['SubscriberEnquiryDate'].date() if hasattr(record['SubscriberEnquiryDate'], 'date') else record['SubscriberEnquiryDate']
                                subscriber_enquiry_date_str = date_only.strftime('%Y-%m-%d')
                        
                        details_viewed_date_str = ""
                        if record['DetailsViewedDate']:
                            if isinstance(record['DetailsViewedDate'], str):
                                details_viewed_date_str = record['DetailsViewedDate']
                            else:
                                date_only = record['DetailsViewedDate'].date() if hasattr(record['DetailsViewedDate'], 'date') else record['DetailsViewedDate']
                                details_viewed_date_str = date_only.strftime('%Y-%m-%d')
                        
                        # Build row data dictionary
                        row_data = {
                            2: current_row - serial_number_base,  # Serial Number
                            3: "",  # Branch ID
                            4: "",  # Unique Tracking Number
                            5: record['SubscriberName'],
                            7: record['SystemUser'] if record['SystemUser'] else "",
                            10: subscriber_enquiry_date_str,
                            11: record['ProductName'],
                            12: details_viewed_date_str,
                            15: record['SearchOutput'] if record['SearchOutput'] else ""
                        }
                        
                        all_rows_data.append((current_row, row_data, subscriber_enquiry_date_str, details_viewed_date_str))
                    
                    # OPTIMIZATION: Use batch_write_rows with template formatting
                    template_data_row = 36
                    batch_write_rows(
                        current_sheet, 
                        current_row_offset,
                        [row_data for _, row_data, _, _ in all_rows_data],
                        template_row=template_data_row
                    )
                    
                    # Set date number formats (template handles styling)
                    print(f"      Applying date formats...")
                    for current_row, _, subscriber_enquiry_date_str, details_viewed_date_str in all_rows_data:
                        if subscriber_enquiry_date_str:
                            current_sheet.cell(row=current_row, column=10).number_format = '@'
                        if details_viewed_date_str:
                            current_sheet.cell(row=current_row, column=12).number_format = '@'
                    
                    data_write_elapsed = time.time() - data_write_start
                    print(f"      Data rows written in {data_write_elapsed:.2f}s")
                    
                    # Move offset to after this product's data for next product
                    current_row_offset += len(product_records)
                    
                    # Set lastProduct for the next iteration
                    lastProduct = product_name
                    
                    product_elapsed = time.time() - product_start_time
                    print(f"      ✓ Product completed in {product_elapsed:.2f}s")

            write_elapsed = time.time() - write_start
            print(f"   ✓ Data written to Excel in {write_elapsed:.2f}s")

            # Auto-size columns for better readability
            for sheet in wb.worksheets:
                auto_size_columns(sheet)

            # Add 'Generated by: <username>' to the first available merged O-Q row, or row 8 if none, with bold and italic formatting
            add_generated_by(ws, request.user.username, current_row_offset - 1)

            # OPTIMIZED: Save workbook directly to file instead of using buffer
            print(f"\n[5/5] Saving file...")
            save_start = time.time()
            
            month_year = start_date.strftime('%B%Y')
            clean_subscriber = clean_filename(subscriber_id)
            filename = f"{clean_subscriber}_{month_year}_{uuid.uuid4().hex[:8]}.xlsx"
            single_reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'single')
            os.makedirs(single_reports_dir, exist_ok=True)
            file_path = os.path.join(single_reports_dir, filename)
            
            # Save directly to final location and free memory
            wb.save(file_path)
            wb.close()
            del wb
            gc.collect()
            
            save_elapsed = time.time() - save_start
            total_elapsed = time.time() - start_time
            print(f"   ✓ File saved as {filename} in {save_elapsed:.2f}s")
            print(f"\n{'='*60}")
            print(f"✓ REPORT GENERATION COMPLETE")
            print(f"   Total Time: {total_elapsed:.2f}s")
            if include_bills:
                print(f"   Billing Products: {sum(summary_bills.values())}")
            if include_products:
                print(f"   Product Records: {len(product_data)}")
            print(f"{'='*60}\n")
            
            download_url = settings.MEDIA_URL + f'reports/single/{filename}'
            
            # Update report generation status to success
            if 'report_gen' in locals() and report_gen:
                report_gen.status = 'success'
                report_gen.completed_at = timezone.now()
                report_gen.save()
            
            # PERFORMANCE: Log execution time
            execution_time = time.time() - start_time
            logger.info(f"Single report generated in {execution_time:.2f} seconds for {subscriber_id}")
                
            return render(request, 'bulkrep/download_ready.html', {
                'download_url': download_url
            })
            
        except Exception as e:
            error_msg = f"Error generating report: {str(e)}"
            messages.error(request, error_msg)
            if 'report_gen' in locals() and report_gen:
                report_gen.status = 'failed'
                report_gen.error_message = error_msg[:500]  # Truncate to fit in the field
                report_gen.completed_at = timezone.now()
                report_gen.save()
            return render(request, 'bulkrep/single_report.html', context)

    return render(request, 'bulkrep/single_report.html', context)

# Helper function to auto-size columns for better readability
def auto_size_columns(worksheet, min_col=1, max_col=17):
    """
    Auto-size columns in the worksheet to fit their contents with reasonable widths.
    Applies modest adjustments to prevent columns from being too wide.
    Handles merged cells properly, especially for rows 2 and 6.
    """
    # Define minimum and maximum widths for specific columns
    min_widths = {
        # Set minimum widths to prevent columns from being too narrow
        5: 10,  # SubscriberName (E)
        6: 10,  # SubscriberName (F)
        7: 9,  # SystemUser (G)
        8: 10,  # SystemUser (H)
        9: 9,  # SystemUser (I)
        10: 15, # SubscriberEnquiryDate (J)
        11: 20, # ProductName (K)
        12: 5, # DetailsViewedDate (L)
        13: 10, # DetailsViewedDate (M)
        14: 5, # DetailsViewedDate (N)
        # Default minimum width for other columns
        'default': 10
    }


    
    # Define maximum widths for specific columns that tend to get too wide
    max_widths = {
        # DetailsViewedDate (columns L-N, 12-14)
        12: 5, 13: 10, 14: 5,
        # SearchOutput (columns O-Q, 15-17) - increased to allow more content to be visible
        15: 40, 16: 40, 17: 40,
        # Default max width for other columns
        'default': 30
    }
    
    # Store merged ranges for special handling
    merged_ranges = list(worksheet.merged_cells.ranges)
    
    # Special handling for rows 2 and 6 which have specific merged ranges
    row2_merged_range = None
    row6_merged_range = None
    
    for merged_range in merged_ranges:
        if merged_range.min_row == 2 and merged_range.max_row == 2:
            row2_merged_range = merged_range
        elif merged_range.min_row == 6 and merged_range.max_row == 6:
            row6_merged_range = merged_range
    
    for col_idx in range(min_col, max_col + 1):
        # Get the maximum content width in the column
        max_length = 0
        column = worksheet.column_dimensions[get_column_letter(col_idx)]
        
        # Skip columns that are part of the merged ranges for rows 2 and 6
        # This prevents these merged cells from affecting column widths
        skip_column = False
        if row2_merged_range and col_idx >= row2_merged_range.min_col and col_idx <= row2_merged_range.max_col:
            # For row 2 merged range, only consider the first column in the range
            if col_idx > row2_merged_range.min_col:
                skip_column = True
        
        if row6_merged_range and col_idx >= row6_merged_range.min_col and col_idx <= row6_merged_range.max_col:
            # For row 6 merged range, only consider the first column in the range
            if col_idx > row6_merged_range.min_col:
                skip_column = True
        
        if not skip_column:
            # Check all cells in this column
            for row_idx, row in enumerate(worksheet.rows, 1):
                # Skip rows 2 and 6 for all columns except the first column in their merged ranges
                if (row_idx == 2 and row2_merged_range and col_idx != row2_merged_range.min_col) or \
                   (row_idx == 6 and row6_merged_range and col_idx != row6_merged_range.min_col):
                    continue
                
                if len(row) >= col_idx:
                    cell = row[col_idx-1]  # 0-based index
                    if cell.value:
                        # Calculate the approximate width based on content
                        try:
                            cell_length = len(str(cell.value))
                            # Adjust for merged cells
                            is_in_merge = False
                            for merged_range in merged_ranges:
                                if cell.coordinate in merged_range:
                                    # Special handling for different types of merged cells
                                    merge_width = merged_range.max_col - merged_range.min_col + 1
                                    
                                    # For rows 2 and 6, handle specially
                                    if (row_idx == 2 and merged_range == row2_merged_range) or \
                                       (row_idx == 6 and merged_range == row6_merged_range):
                                        # For the first column in the merge, allocate more space
                                        if col_idx == merged_range.min_col:
                                            # Allocate more space to first column but not all
                                            cell_length = cell_length * 0.4  # 40% to first column
                                        else:
                                            # Distribute remaining 60% evenly among other columns
                                            cell_length = (cell_length * 0.6) / (merge_width - 1)
                                    
                                    # Special handling for SearchOutput columns (O-Q, 15-17)
                                    elif 15 <= merged_range.min_col <= 17 and 15 <= merged_range.max_col <= 17:
                                        # For SearchOutput with wrap_text, distribute more evenly
                                        # Give more width to first column for better readability
                                        if col_idx == merged_range.min_col:
                                            cell_length = cell_length * 0.5  # 50% to first column
                                        else:
                                            # Distribute remaining 50% evenly
                                            cell_length = (cell_length * 0.5) / (merge_width - 1)
                                    else:
                                        # For other merged cells, divide by the number of columns
                                        # but ensure a minimum reasonable length
                                        cell_length = max(10, cell_length / merge_width)
                                    
                                    is_in_merge = True
                                    break
                            
                            # Update max_length regardless of merge status
                            # This ensures all cells contribute to column width calculation
                            max_length = max(max_length, cell_length)
                        except:
                            pass
        
        # Set column width with appropriate padding and respect minimum/maximum widths
        if max_length > 0:
            # Add padding based on content type - more padding for text columns
            if col_idx >= 15 and col_idx <= 17:  # SearchOutput columns
                # For SearchOutput columns with wrap_text, use more generous padding
                calculated_width = max_length + 3
            else:
                # For other columns, use standard padding
                calculated_width = max_length + 2
            
            # Apply column-specific minimum and maximum width constraints
            column_min_width = min_widths.get(col_idx, min_widths['default'])
            column_max_width = max_widths.get(col_idx, max_widths['default'])
            
            # Set the width to be between minimum and maximum constraints
            column.width = max(column_min_width, min(calculated_width, column_max_width))
        else:
            # If no content, use the minimum width
            column.width = min_widths.get(col_idx, min_widths['default'])

def copy_row_format(ws, template_row_idx, target_row_idx, max_col=15):
    """
    Copy cell styles (and optionally values) from a template row to a target row.
    Also handles merged cells and maintains their formatting.
    """
    # First, check and unmerge any existing merged cells in the target row
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= target_row_idx <= merged_range.max_row:
            ws.unmerge_cells(str(merged_range))
    
    # Copy format and check for merged cells in template
    template_merged_ranges = []
    for col in range(1, max_col + 1):
        template_cell = ws.cell(row=template_row_idx, column=col)
        target_cell = ws.cell(row=target_row_idx, column=col)
        
        # Copy style and number format
        if hasattr(template_cell, '_style'):
            target_cell._style = copy(template_cell._style)
        if hasattr(template_cell, 'number_format'):
            target_cell.number_format = template_cell.number_format
        if hasattr(template_cell, 'alignment') and template_cell.alignment:
            # Create a new alignment object instead of copying the StyleProxy
            h_align = template_cell.alignment.horizontal if template_cell.alignment.horizontal else 'general'
            v_align = template_cell.alignment.vertical if template_cell.alignment.vertical else 'bottom'
            target_cell.alignment = Alignment(horizontal=h_align, vertical=v_align)
        
        # Check if this cell is part of a merged range in the template
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == template_row_idx and merged_range.min_col <= col <= merged_range.max_col:
                template_merged_ranges.append((
                    merged_range.min_col,
                    merged_range.max_col,
                    merged_range.max_col - merged_range.min_col + 1
                ))
                break
    
    # Recreate merged ranges in the target row
    for min_col, max_col, span in set(template_merged_ranges):
        # Set alignment for all cells in the range before merging
        template_main_cell = ws.cell(row=template_row_idx, column=min_col)
        
        # Get alignment values from template cell
        h_align = 'center'  # Default to center
        v_align = 'center'  # Default to center
        
        if hasattr(template_main_cell, 'alignment') and template_main_cell.alignment:
            if template_main_cell.alignment.horizontal:
                h_align = template_main_cell.alignment.horizontal
            if template_main_cell.alignment.vertical:
                v_align = template_main_cell.alignment.vertical
        
        # Apply alignment to all cells in the range
        for col in range(min_col, max_col + 1):
            target_cell = ws.cell(row=target_row_idx, column=col)
            target_cell.alignment = Alignment(horizontal=h_align, vertical=v_align)
        
        # Now merge the cells
        ws.merge_cells(
            start_row=target_row_idx,
            start_column=min_col,
            end_row=target_row_idx,
            end_column=max_col
        )
        
        # Ensure the merged cell has proper alignment
        merged_cell = ws.cell(row=target_row_idx, column=min_col)
        merged_cell.alignment = Alignment(horizontal=h_align, vertical=v_align)

def copy_merged_and_center(ws, template_ws, template_row_start, template_row_end, target_row_start):
    """
    Copy merged cell structure and center alignment from template header rows to target rows.
    """
    # 1. Copy merged cells
    for m_range in template_ws.merged_cells.ranges:
        if template_row_start <= m_range.min_row <= template_row_end:
            row_offset = target_row_start - template_row_start
            new_min_row = m_range.min_row + row_offset
            new_max_row = m_range.max_row + row_offset
            ws.merge_cells(start_row=new_min_row, start_column=m_range.min_col,
                           end_row=new_max_row, end_column=m_range.max_col)
    # 2. Copy center alignment for header cells
    for row in range(template_row_start, template_row_end + 1):
        for col in range(1, ws.max_column + 1):
            template_cell = template_ws.cell(row=row, column=col)
            if hasattr(template_cell, 'alignment') and template_cell.alignment and template_cell.alignment.horizontal == 'center':
                target_cell = ws.cell(row=target_row_start + (row - template_row_start), column=col)
                # Set center alignment
 
                target_cell.alignment = Alignment(horizontal='center', vertical=template_cell.alignment.vertical)

# Helper function to merge and center header columns
def merge_and_center_header_columns(sheet, start_row, end_row):
    for row in range(start_row, end_row + 1):
        # Set row height to 23.5 for better readability of wrapped text in headers
        sheet.row_dimensions[row].height = 25
        
        # Unmerge any existing merged cells in this row
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.min_row <= row <= merged_range.max_row:
                if ((5 <= merged_range.min_col <= 6) or  # SubscriberName (E-F)
                    (7 <= merged_range.min_col <= 9) or  # SystemUser (G-I)
                    (12 <= merged_range.min_col <= 14) or  # DetailsViewedDate (L-N)
                    (15 <= merged_range.min_col <= 17)):  # SearchOutput (O-Q)
                    sheet.unmerge_cells(str(merged_range))
        
        # Merge SubscriberName (E-F)
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        sheet.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        
        # Merge SystemUser (G-I)
        sheet.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        sheet.cell(row=row, column=7).alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge DetailsViewedDate (L-N)
        sheet.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
        sheet.cell(row=row, column=12).alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge SearchOutput (O-Q) with wrap text
        sheet.merge_cells(start_row=row, start_column=15, end_row=row, end_column=17)
        sheet.cell(row=row, column=15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Helper function to merge and center data row columns
def merge_and_center_data_row(sheet, row):
    """Merge and center columns for a specific row.
    Handles SubscriberName (E-F), SystemUser (G-I), DetailsViewedDate (L-N), and SearchOutput (O-Q).
    Also sets row height to 23.5 for better readability of wrapped text.
    """
    # Set row height to 23.5 for better readability of wrapped text
    sheet.row_dimensions[row].height = 25
    
    # Check for existing merged cells in the target row and unmerge them
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row <= row <= merged_range.max_row:
            # Columns for DetailsViewedDate (L-N), SearchOutput (O-Q), SubscriberName (E-F), SystemUser (G-I)
            if (12 <= merged_range.min_col <= 14) or \
                (15 <= merged_range.min_col <= 17) or \
                (5 <= merged_range.min_col <= 6) or \
                (7 <= merged_range.min_col <= 9):
                sheet.unmerge_cells(str(merged_range))
    
    # Set alignment for SubscriberName (E-F)
    for col in range(5, 7):
        cell = sheet.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    
    # Set alignment for SystemUser (G-I)
    for col in range(7, 10):
        cell = sheet.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set alignment for DetailsViewedDate (L-N)
    for col in range(12, 15):
        cell = sheet.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set alignment for SearchOutput (O-Q) with wrap text
    for col in range(15, 18):
        cell = sheet.cell(row=row, column=col)
        # Use top vertical alignment for better readability with wrapped text
        # Center horizontally but align top vertically for multi-line text
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    # Now merge the cells
    # Merge SubscriberName (E-F)
    sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    # Merge SystemUser (G-I)
    sheet.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    # Merge DetailsViewedDate (L-N)
    sheet.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
    # Merge SearchOutput (O-Q)
    sheet.merge_cells(start_row=row, start_column=15, end_row=row, end_column=17)
    
    # Ensure the merged cells have proper alignment
    # SubscriberName
    cell_E = sheet.cell(row=row, column=5)
    cell_E.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # SystemUser
    cell_G = sheet.cell(row=row, column=7)
    cell_G.alignment = Alignment(horizontal='center', vertical='center')
    
    # DetailsViewedDate
    cell_L = sheet.cell(row=row, column=12)
    cell_L.alignment = Alignment(horizontal='center', vertical='center')
    
    # SearchOutput with wrap text
    cell_O = sheet.cell(row=row, column=15)
    # Use top vertical alignment for better readability with wrapped text
    cell_O.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

# Define product rates mapping
PRODUCT_RATES = {
    'Consumer Snap Check': Decimal('500.00'),
    'Consumer Basic Trace': Decimal('170.00'),
    'Consumer Basic Credit': Decimal('170.00'),
    'Consumer Detailed Credit': Decimal('240.00'),
    'X-Score Consumer Detailed Credit': Decimal('500.00'),
    'Commercial Basic Trace': Decimal('275.00'),
    'Commercial Detailed Credit': Decimal('500.00'),
    'Enquiry Report': Decimal('50.00'),
    'Consumer Dud Cheque': Decimal('0.00'),
    'Commercial Dud Cheque': Decimal('0.00'),
    'Director Basic Report': Decimal('0.00'),
    'Director Detailed Report': Decimal('0.00'),
}

def populate_rate_and_amount(ws, start_row, end_row, subscriber_id):
    """
    Populate rate (columns M-O) and calculate amount (columns P-Q) based on product name (column D).
    
    Args:
        ws: Worksheet object
        start_row: First data row (1-based)
        end_row: Last data row (inclusive)
        subscriber_id: The ID/Name of the subscriber to check for custom rates
    """
    
    for row in range(start_row, end_row + 1):
        # Get product name from column D (4th column)
        product_cell = ws.cell(row=row, column=4)
        product_name_original = str(product_cell.value).strip() if product_cell.value else ""
        product_name_cleaned = product_name_original
        
        # Skip empty product names
        if not product_name_cleaned:
            continue
            
        # Use the safe helper function to get the rate
        rate = get_subscriber_product_rate_safe(
            subscriber_id=subscriber_id,
            product_name=product_name_cleaned,
            default_rate_map=PRODUCT_RATES,
            default_rate_key=product_name_cleaned,
            logger=logger
        )
        
        # Populate rate in columns M-O (merged)
        write_to_cell(ws, row, 13, f"₦{rate:,.2f}")  # Column M
        # Calculate amount (rate * 1) since each row represents one search
        amount = rate * Decimal('1.00')
        # Populate amount in columns P-Q (merged)
        write_to_cell(ws, row, 16, f"₦{amount:,.2f}")  # Column P
       

def add_generated_by(ws, username, last_data_row=None):
    """
    Add 'Report Generated by: <username>' two rows below the last data row (or at row 10 if no data), 
    merging O-Q, Trebuchet MS, bold, italic, centered.
    """
    # If no last_data_row provided, default to row 10
    if last_data_row is None:
        last_data_row = 10
    
    # Add two rows below the last data row
    signature_row = last_data_row + 2
    
    # Merge cells O-Q for the signature line
    ws.merge_cells(start_row=signature_row, start_column=15, end_row=signature_row, end_column=17)
    
    # Set the value and formatting for the signature line
    cell = ws.cell(row=signature_row, column=15)
    cell.value = f"Report Generated by: {username}"
    cell.font = openpyxl.styles.Font(name='Trebuchet MS', bold=True, italic=True, color='FF7F7F7F')
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[signature_row].height = 26
    return signature_row

@login_required
def get_subscriber_product_rate(subscriber_name, product_name, default_rate_key, logger=None):
    """Helper function to get subscriber product rate with better error handling."""
    if logger is None:
        logger = logging.getLogger(__name__)
        
    try:
        # Use filter().first() instead of get() to handle multiple records
        custom_rate_obj = SubscriberProductRate.objects.filter(
            subscriber_name__iexact=subscriber_name, 
            product_name__iexact=product_name
        ).first()
        
        if custom_rate_obj and hasattr(custom_rate_obj, 'rate'):
            # Ensure we're returning a Decimal object
            if isinstance(custom_rate_obj.rate, Decimal):
                rate = custom_rate_obj.rate
            else:
                rate = Decimal(str(custom_rate_obj.rate))
            logger.debug(f"Using custom rate for {subscriber_name} - {product_name}: {rate}")
            return rate
        else:
            raise Exception(f"No custom rate found for {product_name}")
    except Exception as e:
        # Get default rate and ensure it's a Decimal
        default_value = ENQUIRY_RATES.get(default_rate_key, Decimal('0.00'))
        if not isinstance(default_value, Decimal):
            rate = Decimal(str(default_value))
        else:
            rate = default_value
        logger.debug(f"Using default rate for {product_name}: {rate}")
        return rate


def get_subscriber_product_rate_safe(subscriber_name, product_name, default_rate_map, default_rate_key, logger=None):
    """Safe helper function to get subscriber product rate with better error handling for multiple records.
    
    Args:
        subscriber_name: The subscriber name to look up
        product_name: The product name to look up
        default_rate_map: Dictionary of default rates (ENQUIRY_RATES or PRODUCT_RATES)
        default_rate_key: Key to use in the default_rate_map
        logger: Optional logger object
    
    Returns:
        Decimal rate value
    """
    if logger is None:
        logger = logging.getLogger(__name__)
        
    try:
        # Use filter().first() instead of get() to avoid MultipleObjectsReturned
        rate_obj = SubscriberProductRate.objects.filter(
            subscriber_name__iexact=subscriber_name, 
            product_name__iexact=product_name
        ).first()
        
        if rate_obj and hasattr(rate_obj, 'rate'):
            # Ensure we're returning a Decimal object
            if isinstance(rate_obj.rate, Decimal):
                rate = rate_obj.rate
            else:
                rate = Decimal(str(rate_obj.rate))
            logger.debug(f"Using custom rate for {subscriber_name} - {product_name}: {rate}")
            return rate
        else:
            # No custom rate found, use default
            default_value = default_rate_map.get(default_rate_key, Decimal('0.00'))
            # Ensure default value is converted to Decimal
            if not isinstance(default_value, Decimal):
                rate = Decimal(str(default_value))
            else:
                rate = default_value
            logger.debug(f"No custom rate found. Using default rate for {product_name}: {rate}")
            return rate
    except Exception as e:
        # Handle any unexpected errors
        logger.error(f"Error retrieving rate for {subscriber_name} - {product_name}: {e}")
        default_value = default_rate_map.get(default_rate_key, Decimal('0.00'))
        # Ensure default value is converted to Decimal
        if not isinstance(default_value, Decimal):
            rate = Decimal(str(default_value))
        else:
            rate = default_value
        logger.debug(f"Using default rate after error: {rate}")
        return rate

        
@login_required
def bulk_report(request):
    """View for generating bulk reports."""
    # PERFORMANCE: Track execution time
    import time
    start_time = time.time()
    
    # Get the current date range (first day of current month to first day of next month)
    today = date.today()
    first_day_of_month = today.replace(day=1)
    
    # Calculate first day of next month
    if today.month == 12:
        first_day_next_month = date(today.year + 1, 1, 1)
    else:
        first_day_next_month = date(today.year, today.month + 1, 1)
    
    # Get unique subscriber names for the dropdown
    subscribers = Usagereport.objects.values_list('SubscriberName', flat=True).distinct().order_by('SubscriberName')
    
    # Initialize report generation tracking
    report_gen = None
    
    # Initial context with date range and subscribers
    context = {
        'start_date': first_day_of_month.strftime('%Y-%m-%d'),
        'end_date': first_day_next_month.strftime('%Y-%m-%d'),
        'subscribers': [{'id': i, 'name': name} for i, name in enumerate(subscribers, 1)],
    }
    
    if request.method == 'POST':
        subscriber_ids = request.POST.getlist('subscribers')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        include_bills = request.POST.get('include_bills') == 'on'
        include_products = request.POST.get('include_products') == 'on'
        output_format = request.POST.get('output_format', 'excel')  # 'excel' or 'pdf'
        
        # Validate that at least one subscriber is selected
        if not subscriber_ids:
            messages.error(request, 'Please select at least one subscriber.')
            return render(request, 'bulkrep/bulk_report.html', context)
            
        # Get the selected subscriber names
        selected_subscribers = [subscribers[int(id)-1] for id in subscriber_ids if int(id) <= len(subscribers)]

        # Create report generation record at the start
        report_gen = ReportGeneration.objects.create(
            user=request.user,
            generator=request.user.username,
            report_type='bulk',
            status='in_progress',
            subscriber_name=f"{len(selected_subscribers)} Subscribers",
            from_date=start_date_str if start_date_str else None,
            to_date=end_date_str if end_date_str else None
        )
        
        # Print tracking banner
        print(f"\n{'='*60}")
        print(f"BULK REPORT GENERATION STARTED")
        print(f"{'='*60}")
        print(f"User: {request.user.username}")
        print(f"Subscribers: {len(selected_subscribers)}")
        print(f"Date Range: {start_date_str} to {end_date_str}")
        print(f"Include Bills: {include_bills}")
        print(f"Include Products: {include_products}")
        print(f"Output Format: {output_format}")
        print(f"{'='*60}\n")
        
        # Process the form submission

        # Convert date strings to date objects - this is critical for display formatting later
        try:
            # Parse the input date strings
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            # Validate date range
            if start_date > end_date:
                messages.error(request, 'Start date cannot be after end date.')
                return render(request, 'bulkrep/bulk_report.html', context)
                
            # Format the dates for display in the report (DD/MM/YYYY)
            start_date_display = start_date.strftime('%d/%m/%Y')
            end_date_display = end_date.strftime('%d/%m/%Y')
            
        except (ValueError, TypeError) as e:
            messages.error(request, f"Invalid date format: {str(e)}")
            return render(request, 'bulkrep/bulk_report.html', context)

        # Get the selected subscriber names
        subscribers_list = selected_subscribers
        
        # Log the selected subscribers for debugging
        print(f"Selected subscribers: {subscribers_list}")
        
        # Validate that we have subscribers to process
        if not subscribers_list:
            messages.error(request, 'No valid subscribers selected.')
            return render(request, 'bulkrep/bulk_report.html', context)
            
        # Remove any empty strings just in case
        subscribers_list = list(filter(None, set(subscribers_list)))
        
        if not subscribers_list:
            messages.warning(request, f"No subscribers found for the selected criteria between {start_date_display} and {end_date_display}.")
            return render(request, 'bulkrep/bulk_report.html', context)

        # OPTIMIZATION: Fetch all usage data and product rates upfront to eliminate N+1 queries
        print(f"[1/4] Fetching usage data for {len(subscribers_list)} subscribers...")
        fetch_start = time.time()
        
        # Fetch all usage data in one query
        all_usage_data = list(Usagereport.objects.filter(
            DetailsViewedDate__gte=start_date,
            DetailsViewedDate__lte=end_date,
            SubscriberName__in=subscribers_list
        ).values(
            'SubscriberName', 'ProductName', 'SystemUser', 'SearchIdentity', 
            'SubscriberEnquiryDate', 'SearchOutput', 'DetailsViewedDate', 
            'ProductInputed'
        ))
        
        # Group usage data by subscriber for instant lookup
        usage_by_subscriber = {}
        for record in all_usage_data:
            subscriber = record['SubscriberName']
            if subscriber not in usage_by_subscriber:
                usage_by_subscriber[subscriber] = []
            usage_by_subscriber[subscriber].append(record)
        
        # Fetch all custom product rates in one query
        print(f"[2/4] Fetching custom product rates...")
        all_custom_rates = list(SubscriberProductRate.objects.filter(
            subscriber_name__in=subscribers_list
        ).values('subscriber_name', 'product_name', 'rate'))
        
        # Create lookup dictionary for custom rates: {(subscriber, product): rate}
        custom_rates_lookup = {}
        for rate_record in all_custom_rates:
            key = (rate_record['subscriber_name'].lower(), rate_record['product_name'].lower())
            custom_rates_lookup[key] = Decimal(str(rate_record['rate']))
        
        fetch_elapsed = time.time() - fetch_start
        print(f"   ✓ Loaded {len(all_usage_data)} usage records and {len(all_custom_rates)} custom rates in {fetch_elapsed:.2f}s")
        
        # Helper function to get custom rate with fallback to default
        def get_custom_rate(subscriber_name, product_name, default_rate):
            """Get custom rate for subscriber/product with fallback to default rate."""
            key = (subscriber_name.lower(), product_name.lower())
            return custom_rates_lookup.get(key, Decimal(str(default_rate)))
        
        # Generate reports for all selected subscribers
        try:
            print(f"\n[3/4] Generating individual reports...")
            overall_start = time.time()  # Track total generation time
            generation_start = time.time()
            
            # Handle PDF generation
            if output_format == 'pdf':
                print(f"   Generating PDF reports...")
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    processed_subscribers = []
                    
                    for idx, subscriber_name in enumerate(subscribers_list, 1):
                        try:
                            print(f"   [{idx}/{len(subscribers_list)}] Generating PDF for {subscriber_name}...")
                            sub_start = time.time()
                            
                            # Use pre-fetched data
                            subscriber_usage_data = usage_by_subscriber.get(subscriber_name, [])
                            
                            if not subscriber_usage_data and include_products:
                                messages.warning(request, f"No data found for subscriber {subscriber_name} between {start_date_display} and {end_date_display}.")
                                continue
                            
                            # Initialize summary bills
                            if include_bills:
                                summary_bills = {
                                    'consumer_snap_check': 0,
                                    'consumer_basic_trace': 0,
                                    'consumer_basic_credit': 0,
                                    'consumer_detailed_credit': 0,
                                    'xscore_consumer_detailed_credit': 0,
                                    'commercial_basic_trace': 0,
                                    'commercial_detailed_credit': 0,
                                    'enquiry_report': 0,
                                    'consumer_dud_cheque': 0,
                                    'commercial_dud_cheque': 0,
                                    'director_basic_report': 0,
                                    'director_detailed_report': 0
                                }
                                
                                for record in subscriber_usage_data:
                                    product_name = record['ProductName']
                                    if 'Snap Check' in product_name:
                                        summary_bills['consumer_snap_check'] += 1
                                    elif 'Basic Trace' in product_name:
                                        summary_bills['consumer_basic_trace'] += 1
                                    elif 'Basic Credit' in product_name:
                                        summary_bills['consumer_basic_credit'] += 1
                                    elif 'X-SCore Consumer Detailed Credit' in product_name:
                                        summary_bills['xscore_consumer_detailed_credit'] += 1
                                    elif 'Consumer Detailed Credit' in product_name:
                                        summary_bills['consumer_detailed_credit'] += 1
                                    elif 'Commercial Basic Trace' in product_name:
                                        summary_bills['commercial_basic_trace'] += 1
                                    elif 'Commercial Detailed Credit' in product_name:
                                        summary_bills['commercial_detailed_credit'] += 1
                                    elif 'Enquiry Report' in product_name:
                                        summary_bills['enquiry_report'] += 1
                                    elif 'Consumer Dud Cheque' in product_name:
                                        summary_bills['consumer_dud_cheque'] += 1
                                    elif 'Commercial Dud Cheque' in product_name:
                                        summary_bills['commercial_dud_cheque'] += 1
                                    elif 'Director Basic Report' in product_name:
                                        summary_bills['director_basic_report'] += 1
                                    elif 'Director Detailed Report' in product_name:
                                        summary_bills['director_detailed_report'] += 1
                                
                                # Get all rates at once using cache (1 query vs 12 queries)
                                cached_rates = get_cached_subscriber_rates(subscriber_name, logger)
                                
                                # Add rates for each product using cached lookup
                                summary_bills['consumer_snap_check_rate'] = get_rate_with_cache(subscriber_name, 'Consumer Snap Check', ENQUIRY_RATES['consumer_snap_check'], cached_rates, logger)
                                summary_bills['consumer_basic_trace_rate'] = get_rate_with_cache(subscriber_name, 'Consumer Basic Trace', ENQUIRY_RATES['consumer_basic_trace'], cached_rates, logger)
                                summary_bills['consumer_basic_credit_rate'] = get_rate_with_cache(subscriber_name, 'Consumer Basic Credit', ENQUIRY_RATES['consumer_basic_credit'], cached_rates, logger)
                                summary_bills['consumer_detailed_credit_rate'] = get_rate_with_cache(subscriber_name, 'Consumer Detailed Credit', ENQUIRY_RATES['consumer_detailed_credit'], cached_rates, logger)
                                summary_bills['xscore_consumer_detailed_credit_rate'] = get_rate_with_cache(subscriber_name, 'X-Score Consumer Detailed Credit', ENQUIRY_RATES['xscore_consumer_detailed_credit'], cached_rates, logger)
                                summary_bills['commercial_basic_trace_rate'] = get_rate_with_cache(subscriber_name, 'Commercial Basic Trace', ENQUIRY_RATES['commercial_basic_trace'], cached_rates, logger)
                                summary_bills['commercial_detailed_credit_rate'] = get_rate_with_cache(subscriber_name, 'Commercial Detailed Credit', ENQUIRY_RATES['commercial_detailed_credit'], cached_rates, logger)
                                summary_bills['enquiry_report_rate'] = get_rate_with_cache(subscriber_name, 'Enquiry Report', ENQUIRY_RATES['enquiry_report'], cached_rates, logger)
                                summary_bills['consumer_dud_cheque_rate'] = get_rate_with_cache(subscriber_name, 'Consumer Dud Cheque', ENQUIRY_RATES['consumer_dud_cheque'], cached_rates, logger)
                                summary_bills['commercial_dud_cheque_rate'] = get_rate_with_cache(subscriber_name, 'Commercial Dud Cheque', ENQUIRY_RATES['commercial_dud_cheque'], cached_rates, logger)
                                summary_bills['director_basic_report_rate'] = get_rate_with_cache(subscriber_name, 'Director Basic Report', ENQUIRY_RATES['director_basic_report'], cached_rates, logger)
                                summary_bills['director_detailed_report_rate'] = get_rate_with_cache(subscriber_name, 'Director Detailed Report', ENQUIRY_RATES['director_detailed_report'], cached_rates, logger)
                            else:
                                summary_bills = {}
                            
                            # Prepare product sections
                            if include_products:
                                product_data = sorted(subscriber_usage_data, key=lambda x: (x['ProductName'], x['DetailsViewedDate']))
                                product_sections = {}
                                for record in product_data:
                                    prod_name = record['ProductName']
                                    if prod_name not in product_sections:
                                        product_sections[prod_name] = []
                                    product_sections[prod_name].append(record)
                            else:
                                product_sections = {}
                            
                            # Generate PDF using PDFReportGenerator
                            from .pdf_generator import PDFReportGenerator
                            
                            pdf_generator = PDFReportGenerator()
                            
                            # Convert product_sections dict to flat list for PDF generator
                            product_data_list = []
                            if include_products and product_sections:
                                for prod_name, records in product_sections.items():
                                    product_data_list.extend(records)
                            
                            pdf_buffer = pdf_generator.generate_report(
                                subscriber_name=subscriber_name,
                                start_date_display=start_date_display,
                                end_date_display=end_date_display,
                                summary_bills=summary_bills if include_bills else None,
                                product_data=product_data_list if include_products else None,
                                username=request.user.username
                            )
                            
                            # Add PDF to zip
                            safe_filename = f"{subscriber_name.replace(' ', '_').replace('/', '_')}_{start_date_str}_to_{end_date_str}.pdf"
                            zip_file.writestr(safe_filename, pdf_buffer.getvalue())
                            processed_subscribers.append(subscriber_name)
                            
                            # Free memory immediately to prevent accumulation
                            del pdf_buffer
                            del subscriber_usage_data
                            del product_data_list
                            gc.collect()
                            
                            sub_end = time.time()
                            print(f"      ✓ Generated in {sub_end - sub_start:.2f}s")
                            
                        except Exception as e:
                            logger.error(f"Error generating PDF report for {subscriber_name}: {str(e)}")
                            messages.error(request, f"Error generating PDF report for {subscriber_name}: {str(e)}")
                            continue
                
                generation_end = time.time()
                print(f"   Generated {len(processed_subscribers)} PDF reports in {generation_end - generation_start:.2f}s")
                
                if not processed_subscribers:
                    messages.error(request, "No reports were generated.")
                    report_gen.status = 'failed'
                    report_gen.error_message = "No reports were generated"
                    report_gen.save()
                    return render(request, 'bulkrep/bulk_report.html', context)
                
                # Save zip file to disk (same as Excel bulk reports)
                zip_buffer.seek(0)
                month_year = start_date.strftime('%B%Y')
                zip_filename = f"bulk_pdf_reports_{month_year}_{uuid.uuid4().hex[:8]}.zip"
                bulk_reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'bulk')
                os.makedirs(bulk_reports_dir, exist_ok=True)
                zip_path = os.path.join(bulk_reports_dir, zip_filename)
                
                with open(zip_path, 'wb') as f:
                    f.write(zip_buffer.read())
                
                download_url = settings.MEDIA_URL + f'reports/bulk/{zip_filename}'
                
                # Update report generation record
                report_gen.status = 'completed'
                report_gen.save()
                
                # Print completion banner
                total_time = time.time() - overall_start
                print(f"\n{'='*60}")
                print(f"BULK PDF REPORT GENERATION COMPLETED")
                print(f"{'='*60}")
                print(f"Total Reports: {len(processed_subscribers)}")
                print(f"Total Time: {total_time:.2f}s")
                print(f"Zip File: {zip_filename}")
                print(f"{'='*60}\n")
                
                # Redirect to download page (same as single report and Excel bulk)
                success_msg = f"Successfully generated PDF reports for {len(processed_subscribers)} subscribers."
                messages.success(request, success_msg)
                return render(request, 'bulkrep/download_ready.html', {
                    'download_url': download_url
                })
            
            # Handle Excel generation
            else:
                zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                processed_subscribers = []
                
                for idx, subscriber_name in enumerate(subscribers_list, 1):
                    try:
                        # Log which subscriber we're processing
                        print(f"   [{idx}/{len(subscribers_list)}] Processing {subscriber_name}...")
                        sub_start = time.time()
                        
                        # OPTIMIZED: Use pre-fetched data instead of database queries
                        subscriber_usage_data = usage_by_subscriber.get(subscriber_name, [])
                        
                        if include_bills:
                            # Initialize summary dictionary with all possible keys
                            summary_bills = {
                                'consumer_snap_check': 0,
                                'consumer_basic_trace': 0,
                                'consumer_basic_credit': 0,
                                'consumer_detailed_credit': 0,
                                'xscore_consumer_detailed_credit': 0,
                                'commercial_basic_trace': 0,
                                'commercial_detailed_credit': 0,
                                'enquiry_report': 0,
                                'consumer_dud_cheque': 0,
                                'commercial_dud_cheque': 0,
                                'director_basic_report': 0,
                                'director_detailed_report': 0
                            }
                            
                            # Count each product type using Python filtering (much faster than DB queries)
                            for record in subscriber_usage_data:
                                product_name = record['ProductName']
                                if 'Snap Check' in product_name:
                                    summary_bills['consumer_snap_check'] += 1
                                elif 'Basic Trace' in product_name:
                                    summary_bills['consumer_basic_trace'] += 1
                                elif 'Basic Credit' in product_name:
                                    summary_bills['consumer_basic_credit'] += 1
                                elif 'X-SCore Consumer Detailed Credit' in product_name:
                                    summary_bills['xscore_consumer_detailed_credit'] += 1
                                elif 'Consumer Detailed Credit' in product_name:
                                    summary_bills['consumer_detailed_credit'] += 1
                                elif 'Commercial Basic Trace' in product_name:
                                    summary_bills['commercial_basic_trace'] += 1
                                elif 'Commercial Detailed Credit' in product_name:
                                    summary_bills['commercial_detailed_credit'] += 1
                                elif 'Enquiry Report' in product_name:
                                    summary_bills['enquiry_report'] += 1
                                elif 'Consumer Dud Cheque' in product_name:
                                    summary_bills['consumer_dud_cheque'] += 1
                                elif 'Commercial Dud Cheque' in product_name:
                                    summary_bills['commercial_dud_cheque'] += 1
                                elif 'Director Basic Report' in product_name:
                                    summary_bills['director_basic_report'] += 1
                                elif 'Director Detailed Report' in product_name:
                                    summary_bills['director_detailed_report'] += 1
                        else:
                            summary_bills = {}

                        # OPTIMIZED: Use pre-fetched data for product details
                        if include_products:
                            # Sort the data by ProductName and DetailsViewedDate
                            product_data = sorted(subscriber_usage_data, key=lambda x: (x['ProductName'], x['DetailsViewedDate']))
                            
                            # Group by ProductName
                            product_sections = {}
                            for record in product_data:
                                product_name = record['ProductName']
                                if product_name not in product_sections:
                                    product_sections[product_name] = []
                                product_sections[product_name].append(record)
                        else:
                            product_sections = {}
                            product_data = []

                        if not product_data and include_products:
                            messages.warning(request, f"No data found for subscriber {subscriber_name} between {start_date_display} and {end_date_display}.")
                            continue

                        # Generate Excel report using JSON-based builder
                        try:
                            # Prepare dynamic values for the template
                            value_overrides = excel_builder.prepare_dynamic_values(
                                subscriber_name, start_date_display, end_date_display
                            )
                            
                            # Create workbook from JSON structure (NOT from .xlsx template)
                            wb = excel_builder.create_workbook_from_json(value_overrides)
                            ws = wb.active
                        except Exception as e:
                            logger.error(f"Error creating workbook from JSON template: {str(e)}")
                            messages.error(request, f"Error creating workbook from JSON template: {str(e)}")
                            return render(request, 'bulkrep/bulk_report.html', context)
                        try:
                            
                            if include_bills:
                                # Fetch rates for the subscriber and add to summary_bills
                                cached_rates = get_cached_subscriber_rates(subscriber_name, logger)
                                
                                # Add rates to summary_bills dict
                                for key, name in [
                                    ('consumer_snap_check', 'Consumer Snap Check'),
                                    ('consumer_basic_trace', 'Consumer Basic Trace'),
                                    ('consumer_basic_credit', 'Consumer Basic Credit'),
                                    ('consumer_detailed_credit', 'Consumer Detailed Credit'),
                                    ('xscore_consumer_detailed_credit', 'X-Score Consumer Detailed Credit'),
                                    ('commercial_basic_trace', 'Commercial Basic Trace'),
                                    ('commercial_detailed_credit', 'Commercial Detailed Credit'),
                                    ('enquiry_report', 'Enquiry Report'),
                                    ('consumer_dud_cheque', 'Consumer Dud Cheque'),
                                    ('commercial_dud_cheque', 'Commercial Dud Cheque'),
                                    ('director_basic_report', 'Director Basic Report'),
                                    ('director_detailed_report', 'Director Detailed Report'),
                                ]:
                                    default_rate = ENQUIRY_RATES.get(key, Decimal('0.00'))
                                    summary_bills[f'{key}_rate'] = get_rate_with_cache(
                                        subscriber_name, name, default_rate, cached_rates, logger
                                    )
                                
                                # Use excel_builder to write billing data (handles all billing rows with styling)
                                excel_builder.write_billing_data(ws, summary_bills, subscriber_name, logger)
                            
                            if include_products:
                                start_row_offset = 36  # Initial start row for product sections on Sheet1
                                current_sheet = ws
                                sheet2 = wb["Sheet2"] if "Sheet2" in wb.sheetnames else None
                                
                                # Sort product_sections by product_name
                                sorted_product_names = sorted(product_sections.keys())
                                
                                # First, find the original product name cell in the template (rows 32-35)
                                product_name_cell = None
                                for row in range(32, 36):  # Rows 32-35 (inclusive)
                                    for col in range(1, 16):  # Assuming columns A-O are important
                                        cell_value = ws.cell(row=row, column=col).value
                                        if cell_value and "product" in str(cell_value).lower():
                                            product_name_cell = (row, col)
                                            break
                                    if product_name_cell:
                                        break
                                        
                                # Save the template header structure (rows 32-35) for subsequent products
                                header_template = []
                                header_rows = (32, 35)  # Range of rows to copy for the header template
                                for row in range(header_rows[0], header_rows[1] + 1):
                                    row_data = []
                                    for col in range(1, 16):  # Assuming columns A-O are important
                                        cell = ws.cell(row=row, column=col)
                                        # Store cell value and position only - we'll copy styles directly later
                                        cell_info = {
                                            'value': cell.value,
                                            'position': (row, col)
                                        }
                                        # Check if cell is part of merged range
                                        for m_range in ws.merged_cells.ranges:
                                            if (row, col) == (m_range.min_row, m_range.min_col):
                                                cell_info['merged'] = (m_range.max_row - m_range.min_row + 1, 
                                                                    m_range.max_col - m_range.min_col + 1)
                                                break
                                        row_data.append(cell_info)
                                    header_template.append(row_data)
                                
                                # Set data row start - will be used for first product                
                                data_start_row = 36  # Initial start for data rows after template header
                                lastProduct = ""
                                
                                # For each product, use appropriate header section
                                for product_idx, product_name in enumerate(sorted_product_names):
                                    product_records = product_sections[product_name]
                                    
                                    if product_idx == 0:
                                        # For first product, use the existing template header
                                        if product_name_cell:
                                            row, col = product_name_cell
                                            # Replace "Product Name" with actual first product name in template
                                            safe_cell_assignment(ws, row, col, product_name)
                                        current_sheet = ws
                                        current_row_offset = data_start_row  # Start data at row 36
                                        serial_number_base = data_start_row - 1
                                    else:
                                        # Add space between different products (extra spacing)
                                        current_row_offset += 4  # Add significant spacing between products
                                            
                                        # Create new header for subsequent products
                                        header_start_row = current_row_offset
                                        
                                        # Clone the header section for this product
                                        for template_row_idx, template_row in enumerate(header_template):
                                            target_row = header_start_row + template_row_idx
                                        
                                            # Ensure we're not past row limits
                                            if current_sheet == ws and target_row > 1000000 and sheet2:
                                                current_sheet = sheet2
                                                # Reset for Sheet2
                                                header_start_row = 13
                                                target_row = header_start_row + template_row_idx
                                        
                                            # Unmerge any existing merged cells in the target area
                                            for m_range in list(current_sheet.merged_cells.ranges):
                                                if m_range.min_row <= target_row <= m_range.max_row:
                                                    current_sheet.unmerge_cells(
                                                        start_row=m_range.min_row, 
                                                        start_column=m_range.min_col,
                                                        end_row=m_range.max_row, 
                                                        end_column=m_range.max_col
                                                    )
                                        
                                            # Copy each cell from the template to the target area
                                            for col_idx, cell_info in enumerate(template_row):
                                                target_col = col_idx + 1
                                                target_cell = current_sheet.cell(row=target_row, column=target_col)
                                                
                                                # Copy styles directly from the original cell
                                                original_row, original_col = cell_info['position']
                                                original_cell = ws.cell(row=original_row, column=original_col)
                                                
                                                # Copy cell format using openpyxl's built-in method
                                                target_cell._style = copy(original_cell._style)
                                                
                                                # Set value using safe_cell_assignment to handle merged cells properly
                                                if template_row_idx == product_name_cell[0] - header_rows[0] and \
                                                col_idx == product_name_cell[1] - 1:
                                                    # For the product name cell, use the safe assignment method
                                                    safe_cell_assignment(current_sheet, target_row, target_col, product_name)
                                                elif cell_info['value'] is not None:
                                                    # For other cells with values, use the safe assignment method
                                                    safe_cell_assignment(current_sheet, target_row, target_col, cell_info['value'])
                                                
                                                # Recreate merged cells
                                                if 'merged' in cell_info:
                                                    rows, cols = cell_info['merged']
                                                    current_sheet.merge_cells(
                                                        start_row=target_row, 
                                                        start_column=target_col,
                                                        end_row=target_row + rows - 1, 
                                                        end_column=target_col + cols - 1
                                                    )
                                        
                                        # Update data start row to be after this new header
                                        current_row_offset = header_start_row + (header_rows[1] - header_rows[0] + 1)
                                        serial_number_base = current_row_offset - 1
                                        
                                        # Add header for the data section
                                        safe_cell_assignment(current_sheet, current_row_offset - 1, 4, "Unique Tracking Number")

                                    # Process data records for this product - OPTIMIZED
                                    # OPTIMIZATION: Build all row data first, then batch write
                                    all_rows_data = []
                                    for record_idx, record in enumerate(product_records):
                                        current_row = current_row_offset + record_idx
                                        
                                        # Switch to Sheet2 when reaching max row (like VBA)
                                        if current_row > 1000000 and sheet2 and current_sheet != sheet2:
                                            current_sheet = sheet2
                                            current_row_offset = 13  # Reset row offset for Sheet2 as per VBA
                                            serial_number_base = 12  # Reset serial number base for Sheet2
                                            current_row = current_row_offset + record_idx  # Recalculate current_row for Sheet2
                                        
                                        # Format dates
                                        subscriber_enquiry_date_str = ""
                                        if record['SubscriberEnquiryDate']:
                                            if isinstance(record['SubscriberEnquiryDate'], str):
                                                subscriber_enquiry_date_str = record['SubscriberEnquiryDate']
                                            else:
                                                date_only = record['SubscriberEnquiryDate'].date() if hasattr(record['SubscriberEnquiryDate'], 'date') else record['SubscriberEnquiryDate']
                                                subscriber_enquiry_date_str = date_only.strftime('%Y-%m-%d')
                                        
                                        details_viewed_date_str = ""
                                        if record['DetailsViewedDate']:
                                            if isinstance(record['DetailsViewedDate'], str):
                                                details_viewed_date_str = record['DetailsViewedDate']
                                            else:
                                                date_only = record['DetailsViewedDate'].date() if hasattr(record['DetailsViewedDate'], 'date') else record['DetailsViewedDate']
                                                details_viewed_date_str = date_only.strftime('%Y-%m-%d')
                                        
                                        # Build row data dictionary
                                        row_data = {
                                            2: current_row - serial_number_base,  # Serial Number
                                            3: "",  # Branch ID
                                            4: "",  # Unique Tracking Number
                                            5: record['SubscriberName'],
                                            7: record['SystemUser'] if record['SystemUser'] else "",
                                            10: subscriber_enquiry_date_str,
                                            11: record['ProductName'],
                                            12: details_viewed_date_str,
                                            15: record['SearchOutput'] if record['SearchOutput'] else ""
                                        }
                                        
                                        all_rows_data.append((current_row, row_data, subscriber_enquiry_date_str, details_viewed_date_str))
                                    
                                    # OPTIMIZATION: Use batch_write_rows with template formatting
                                    template_data_row = 36
                                    batch_write_rows(
                                        current_sheet, 
                                        current_row_offset,
                                        [row_data for _, row_data, _, _ in all_rows_data],
                                        template_row=template_data_row
                                    )
                                    
                                    # Set date number formats (template handles styling)
                                    for current_row, _, subscriber_enquiry_date_str, details_viewed_date_str in all_rows_data:
                                        if subscriber_enquiry_date_str:
                                            current_sheet.cell(row=current_row, column=10).number_format = '@'
                                        if details_viewed_date_str:
                                            current_sheet.cell(row=current_row, column=12).number_format = '@'
                                    
                                    # Move offset to after this product's data for next product
                                    current_row_offset += len(product_records)
                                    
                                    # Set lastProduct for the next iteration
                                    lastProduct = product_name

                            # Auto-size columns for better readability
                            for sheet in wb.worksheets:
                                auto_size_columns(sheet)

                            add_generated_by(ws, request.user.username, current_row_offset - 1)

                        except Exception as e:
                            logger.error(f"Error generating report data: {str(e)}")
                            messages.error(request, f"Error generating report data: {str(e)}")
                            return render(request, 'bulkrep/bulk_report.html', context)
                        
                        # OPTIMIZED: Write to temp file instead of memory buffer
                        month_year = start_date.strftime('%B%Y')
                        clean_subscriber = clean_filename(subscriber_name)
                        filename = f"{clean_subscriber}_{month_year}.xlsx"
                        
                        # Write Excel to temporary file and add to zip
                        temp_excel_path = write_excel_to_temp_file(wb, clean_subscriber)
                        try:
                            with open(temp_excel_path, 'rb') as temp_file:
                                zip_file.writestr(filename, temp_file.read())
                        finally:
                            # Clean up temporary file
                            if os.path.exists(temp_excel_path):
                                os.unlink(temp_excel_path)
                        
                        # Add to processed subscribers list
                        processed_subscribers.append(subscriber_name)
                        
                        sub_elapsed = time.time() - sub_start
                        print(f"      ✓ Completed in {sub_elapsed:.2f}s")
                        
                    except Exception as e:
                        print(f"      ✗ Error: {str(e)}")
                        error_msg = f"Skipped report for {subscriber_name} due to error: {str(e)}"
                        messages.warning(request, error_msg)
                        if 'report_gen' in locals() and report_gen:
                            report_gen.status = 'failed'
                            report_gen.error_message = error_msg[:500]  # Truncate to fit in the field
                            report_gen.completed_at = timezone.now()
                            report_gen.save()
                        continue
            
            # Return the zip file if any reports were generated
            if processed_subscribers:
                generation_elapsed = time.time() - generation_start
                print(f"   ✓ All reports generated in {generation_elapsed:.2f}s")
                
                print(f"\n[4/4] Creating zip archive...")
                zip_start = time.time()
                
                zip_buffer.seek(0)
                month_year = start_date.strftime('%B%Y')
                zip_filename = f"all_subscriber_reports_{month_year}_{uuid.uuid4().hex[:8]}.zip"
                bulk_reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'bulk')
                os.makedirs(bulk_reports_dir, exist_ok=True)
                zip_path = os.path.join(bulk_reports_dir, zip_filename)
                with open(zip_path, 'wb') as f:
                    f.write(zip_buffer.read())
                download_url = settings.MEDIA_URL + f'reports/bulk/{zip_filename}'
                
                zip_elapsed = time.time() - zip_start
                total_elapsed = time.time() - start_time
                print(f"   ✓ Zip file created in {zip_elapsed:.2f}s")
                print(f"\n{'='*60}")
                print(f"✓ BULK REPORT GENERATION COMPLETE")
                print(f"   Total Time: {total_elapsed:.2f}s")
                print(f"   Subscribers Processed: {len(processed_subscribers)}/{len(subscribers_list)}")
                print(f"   Total Records: {len(all_usage_data)}")
                print(f"   Zip File: {zip_filename}")
                print(f"{'='*60}\n")
                
                # Update report generation status to success
                if 'report_gen' in locals() and report_gen:
                    report_gen.status = 'success'
                    report_gen.completed_at = timezone.now()
                    report_gen.save()
                
                # PERFORMANCE: Log execution time
                execution_time = time.time() - start_time
                logger.info(f"Bulk report generated in {execution_time:.2f} seconds for {len(processed_subscribers)} subscribers")
                
                success_msg = f"Successfully generated reports for {len(processed_subscribers)} out of {len(subscribers_list)} subscribers for the period {start_date_display} to {end_date_display}."
                messages.success(request, success_msg)
                return render(request, 'bulkrep/download_ready.html', {
                    'download_url': download_url
                })
            else:
                messages.warning(request, "No reports were generated. Please check the data or try different criteria.")
                return render(request, 'bulkrep/bulk_report.html', context)
            
        except Exception as e:
            messages.error(request, f"Error generating bulk reports: {str(e)}")
            return render(request, 'bulkrep/bulk_report.html', context)
    
    # For GET requests, render the bulk report form
    return render(request, 'bulkrep/bulk_report.html', context)


# Dashboard Views


@login_required
def dashboard(request):
    """Main dashboard view"""
    if request.user.is_authenticated and not request.user.is_superuser:
        has_assignments = Subscriber.objects.filter(managed_by=request.user).exists()
        if not has_assignments:
            messages.info(request, 'Please select the subscribers you manage.')
            return redirect('bulkrep:subscriber_selection')
    context = {
        'title': '📊 Usage Analytics Dashboard'
    }
    return render(request, 'bulkrep/dashboard.html', context)

def update_subscriber_status(subscriber):
    """
    Updates the subscriber status based on the last usage date.
    If last usage > 2 months ago, status is inactive.
    """
    try:
        last_usage = Usagereport.objects.filter(SubscriberName=subscriber.name).order_by('-DetailsViewedDate').first()
        if last_usage and last_usage.DetailsViewedDate:
            last_date = last_usage.DetailsViewedDate
            
            # Handle different date formats
            if isinstance(last_date, str):
                # Parse string date - try different formats
                date_formats = [
                    '%Y-%m-%d %H:%M:%S.%f',  # 2025-06-21 00:00:00.000
                    '%Y-%m-%d %H:%M:%S',      # 2025-06-21 00:00:00
                    '%Y-%m-%d',                # 2025-06-21
                ]
                for fmt in date_formats:
                    try:
                        last_date = datetime.strptime(last_date, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    # If none of the formats worked
                    logger.error(f"Could not parse date string: {last_date}")
                    return False
            elif isinstance(last_date, datetime):
                last_date = last_date.date()
            # If it's already a date object, use it as is
            
            two_months_ago = date.today() - timedelta(days=60)
            
            new_status = 'active' if last_date >= two_months_ago else 'inactive'
            
            if subscriber.status != new_status:
                subscriber.status = new_status
                subscriber.save()
                return True
    except Exception as e:
        logger.error(f"Error updating status for {subscriber.name}: {e}")
    return False

@login_required
def my_subscribers(request):
    qs = Subscriber.objects.filter(managed_by=request.user).order_by('name')
    # Update statuses for all subscribers
    for sub in qs:
        update_subscriber_status(sub)
    return render(request, 'bulkrep/manage_subscribers.html', {'subscribers': qs})

class SubscriberSelectionForm(forms.Form):
    selected_names = forms.MultipleChoiceField(
        required=True,
        widget=forms.CheckboxSelectMultiple,
        choices=[],
        label='Select subscribers you manage'
    )

@login_required
def subscriber_selection(request):
    # Managers can revisit selection anytime; superusers can access too

    # Compute available subscriber names from usage, excluding already assigned to someone else
    all_names = list(Usagereport.objects.values_list('SubscriberName', flat=True).distinct())
    taken_names = set(Subscriber.objects.exclude(managed_by__isnull=True).values_list('name', flat=True))
    available = [n for n in all_names if n and n not in taken_names]
    available.sort()

    form = SubscriberSelectionForm()
    form.fields['selected_names'].choices = [(n, n) for n in available]

    if request.method == 'POST':
        form = SubscriberSelectionForm(request.POST)
        form.fields['selected_names'].choices = [(n, n) for n in available]
        if form.is_valid():
            selected = form.cleaned_data['selected_names']
            created_count = 0
            for name in selected:
                obj, _ = Subscriber.objects.get_or_create(name=name)
                # Skip if someone else claimed meanwhile
                if obj.managed_by and obj.managed_by_id != request.user.id:
                    continue
                obj.managed_by = request.user
                if not obj.status:
                    obj.status = 'pending'
                obj.save()
                created_count += 1
        
            # Invalidate dashboard cache for this user
            invalidate_user_dashboard_cache(request.user.id)
            messages.success(request, f'Saved {created_count} subscribers.')
            return redirect('bulkrep:my_subscribers')

    return render(request, 'bulkrep/subscriber_selection.html', {'form': form})

class SubscriberForm(forms.ModelForm):
    class Meta:
        model = Subscriber
        fields = ['name', 'contact_person', 'email', 'phone_number']
        widgets = {
            'name': forms.Select(attrs={'class': 'form-select'}),
            'contact_person': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Contact person'}),
            'email': forms.EmailInput(attrs={'class': 'form-control', 'placeholder': 'Email address'}),
            'phone_number': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Phone number', 'type': 'tel'}),
        }
        labels = {
            'name': 'Subscriber',
        }

@login_required
def subscriber_create(request):
    available_names = list(Usagereport.objects.values_list('SubscriberName', flat=True).distinct())
    assigned_names = set(Subscriber.objects.values_list('name', flat=True))
    choices = [(n, n) for n in available_names if n and n not in assigned_names]
    if request.method == 'POST':
        form = SubscriberForm(request.POST)
        form.fields['name'] = forms.ChoiceField(choices=[('', 'Select Subscriber')] + choices)
        form.fields['name'].widget = forms.Select(attrs={'class': 'form-select'})
        if form.is_valid():
            obj = form.save(commit=False)
            obj.managed_by = request.user
            obj.save()
            # Update status based on usage
            update_subscriber_status(obj)
            # Invalidate dashboard cache for this user
            invalidate_user_dashboard_cache(request.user.id)
            return redirect('bulkrep:my_subscribers')
    else:
        form = SubscriberForm()
        form.fields['name'] = forms.ChoiceField(choices=[('', 'Select Subscriber')] + choices)
        form.fields['name'].widget = forms.Select(attrs={'class': 'form-select'})
    return render(request, 'bulkrep/subscriber_form.html', {'form': form, 'title': 'Add Subscriber'})

@login_required
def subscriber_update(request, pk):
    obj = Subscriber.objects.get(pk=pk)
    
    # Update status based on usage before loading form
    update_subscriber_status(obj)
    # Refresh from DB to get updated status
    obj.refresh_from_db()
    
    if not (request.user.is_superuser or obj.managed_by_id == request.user.id):
        return HttpResponse(status=403)
    if request.method == 'POST':
        form = SubscriberForm(request.POST, instance=obj)
        if 'name' in form.fields:
            form.fields['name'].widget = forms.Select(attrs={'class': 'form-select', 'disabled': 'disabled'})
            form.fields['name'].choices = [(obj.name, obj.name)]
            form.fields['name'].disabled = True
             
        if form.is_valid():
            saved_obj = form.save()
            # Update status after saving other fields
            update_subscriber_status(saved_obj)
            # Invalidate dashboard cache for this user
            invalidate_user_dashboard_cache(request.user.id)
            return redirect('bulkrep:my_subscribers')
    else:
        form = SubscriberForm(instance=obj)
        if 'name' in form.fields:
            form.fields['name'].widget = forms.Select(attrs={'class': 'form-select', 'disabled': 'disabled'})
            form.fields['name'].choices = [(obj.name, obj.name)]
            form.fields['name'].disabled = True
    return render(request, 'bulkrep/subscriber_form.html', {'form': form, 'title': 'Edit Subscriber', 'subscriber': obj})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_reassign_subscribers(request):
    # Only show subscribers that have a manager assigned (hide unassigned subscribers)
    qs = Subscriber.objects.select_related('managed_by').filter(managed_by__isnull=False).order_by('name')
    managers = list(Subscriber._meta.get_field('managed_by').remote_field.model.objects.filter(is_superuser=False))
    if request.method == 'POST':
        sid = request.POST.get('subscriber_id')
        uid = request.POST.get('user_id')
        if sid:
            s = Subscriber.objects.get(pk=int(sid))
            s.managed_by_id = int(uid) if uid else None
            s.save()
            # Invalidate cache for the affected user
            if s.managed_by_id:
                invalidate_user_dashboard_cache(s.managed_by_id)
            return redirect('bulkrep:admin_reassign_subscribers')
    return render(request, 'bulkrep/admin_reassign_subscribers.html', {'subscribers': qs, 'managers': managers})

def invalidate_user_dashboard_cache(user_id):
    """
    Invalidate all dashboard cache entries for a specific user.
    This is called when a manager's subscriber assignments change.
    """
    try:
        # Clear all cache keys that contain this user's ID
        # Since we can't easily iterate cache keys, we'll use a cache version approach
        cache_version_key = f"dashboard_cache_version_user_{user_id}"
        current_version = cache.get(cache_version_key, 0)
        cache.set(cache_version_key, current_version + 1, timeout=None)
    except Exception as e:
        logger.error(f"Error invalidating cache for user {user_id}: {str(e)}")


@login_required
def dashboard_api(request):
    """API endpoint for dashboard data with caching and consistent filtering."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=403)

    try:
        # --- Date Handling ---
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            today = timezone.now().date()
            start_date = today.replace(day=1)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            today = timezone.now().date()
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1)

        # --- Filter Handling ---
        subscriber_filter = request.GET.get('subscriber_filter')
        if subscriber_filter and subscriber_filter.lower().strip() in ['all', 'null', '']:
            subscriber_filter = None
        
        # --- Special View Handling ---
        if request.GET.get('three_month_view', 'false').lower() == 'true':
            data = {'three_month_usage': get_three_month_rolling_usage(subscriber_filter)}
            return JsonResponse(data)

        # --- Cache Key Generation ---
        # Include cache version for this user to allow selective invalidation
        cache_version_key = f"dashboard_cache_version_user_{request.user.id if not request.user.is_superuser else 'admin'}"
        cache_version = cache.get(cache_version_key, 0)
        
        cache_params = {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'subscriber_filter': subscriber_filter or 'all',
            'user_id': request.user.id if not request.user.is_superuser else 'admin',  # Include user for managers
            'version': cache_version,  # Include version for cache invalidation
        }
        cache_key_string = '|'.join([f"{k}:{v}" for k, v in sorted(cache_params.items())])
        cache_key = f"dashboard_api_v8_{hashlib.md5(cache_key_string.encode()).hexdigest()}" # Changed to v8 with active_subscribers

        cached_data = cache.get(cache_key)
        if cached_data:
            print(f"CACHE HIT - active_subscribers in cache: {cached_data.get('active_subscribers', 'NOT FOUND')}")
            return JsonResponse(cached_data)

        # --- UNIFIED DATA PAYLOAD ---
        data = {
            'total_subscribers': get_total_subscribers(subscriber_filter, request.user),
            'active_subscribers': get_active_subscribers(start_date, end_date, subscriber_filter, request.user),
            'total_usage_entries': get_total_usage_entries(start_date, end_date, subscriber_filter, request.user),

            'top_subscriber': get_top_subscriber(start_date, end_date, subscriber_filter, request.user),
            'top_subscribers': get_top_subscribers_by_usage_filtered(start_date, end_date, subscriber_filter, 10, request.user),
            'top_products': get_top_products_by_frequency_filtered(start_date, end_date, None, subscriber_filter, 10, request.user),
            'all_products_by_frequency': get_all_products_by_frequency(start_date, end_date, subscriber_filter, request.user),
            'all_subscribers_by_usage': get_all_subscribers_by_usage(start_date, end_date, subscriber_filter, request.user),
            'key_subscribers_list': get_key_subscribers_list() if request.user.is_superuser else [],
            'usage_trends': get_usage_trends_filtered(start_date, end_date, None, subscriber_filter, request.user),
            'new_subscribers': get_new_subscribers_trend_filtered(start_date, end_date, None, subscriber_filter, request.user),
            'retention_rate': get_retention_rate(start_date, end_date, subscriber_filter, request.user),
            'engagement_rate': get_engagement_rate(start_date, end_date, subscriber_filter, request.user),
            'highest_product_by_transaction': get_highest_product_by_transaction(start_date, end_date, subscriber_filter, request.user),

            'unique_products': get_unique_products(subscriber_filter, request.user),
            'unique_subscribers': get_unique_subscribers(request.user),
            'daily_comparison': get_daily_comparison(subscriber_filter, request.user),
            'week_comparison': get_week_comparison(subscriber_filter, request.user),
            'month_comparison': get_month_comparison(subscriber_filter, request.user),
            # 'same_day_comparison': get_same_day_comparison(subscriber_filter, request.user),  # Commented out - not needed for now
            'churn_data': get_churn_data_filtered(start_date, end_date, None, subscriber_filter, request.user),
            'is_admin': request.user.is_superuser,  # Flag for frontend to show/hide admin-only options
        }

        print(f"CACHE MISS - Sending active_subscribers: {data.get('active_subscribers')}")
        cache.set(cache_key, data, 300)
        return JsonResponse(data)

    except Exception as e:
        import traceback
        logger.error(f"Dashboard API error: {traceback.format_exc()}")
        return JsonResponse({'error': 'An error occurred while loading dashboard data.'}, status=500)


# MODIFIED: This function is now correct
# def get_total_subscribers(start_date, end_date, subscriber_filter=None, user=None):
#     """Get total number of unique subscribers for date range, filtered by subscriber if provided"""
#     query = Usagereport.objects.filter(
#         DetailsViewedDate__range=[start_date, end_date]
#     )
#     if user and not user.is_superuser:
#         assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
#         query = query.filter(SubscriberName__in=assigned)
#     if subscriber_filter == 'key_subscribers':
#         query = query.filter(SubscriberName__in=get_key_subscribers_list())
#         return query.values('SubscriberName').distinct().count()
#     elif subscriber_filter and subscriber_filter != 'all':
#         # Check if the single subscriber has any usage in the period
#         return 1 if query.filter(SubscriberName=subscriber_filter).exists() else 0
#     else:
#         return query.values('SubscriberName').distinct().count()
def get_total_subscribers(subscriber_filter=None, user=None):
    """Get total count of all distinct subscribers in the portfolio (not date-filtered).
    
    This counts ALL unique subscriber names that have ever appeared in usage data,
    representing the complete subscriber portfolio.
    """
    try:
        query = Usagereport.objects.all()  # No date filter - all time
        
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
        
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            return 1 if query.filter(SubscriberName=subscriber_filter).exists() else 0
        
        return query.values('SubscriberName').distinct().count()
    except Exception as e:
        logger.error(f"Error getting total subscribers: {str(e)}")
        return 0


def get_active_subscribers(start_date, end_date, subscriber_filter=None, user=None):
    """Get count of subscribers with usage activity in the selected date range.
    
    This counts unique subscribers who have actual usage entries within the
    specified date range, representing active engagement.
    """
    try:
        logger.info(f"get_active_subscribers called: start={start_date}, end={end_date}, filter={subscriber_filter}, user={user}")
        
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
            logger.info(f"Filtered by assigned subscribers: {len(assigned)} assigned")
        
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            exists = query.filter(SubscriberName=subscriber_filter).exists()
            logger.info(f"Single subscriber filter: {subscriber_filter}, exists={exists}")
            return 1 if exists else 0
        
        count = query.values('SubscriberName').distinct().count()
        logger.info(f"Active subscribers count: {count}")
        return count
    except Exception as e:
        logger.error(f"Error getting active subscribers: {str(e)}")
        return 0




# MODIFIED: This function is now correct
def get_total_usage_entries(start_date, end_date, subscriber_filter=None, user=None):
    """Get total usage entries for date range, filtered by subscriber if provided"""
    query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])

    if user and not user.is_superuser:
        assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
        query = query.filter(SubscriberName__in=assigned)
    if subscriber_filter == 'key_subscribers':
        query = query.filter(SubscriberName__in=get_key_subscribers_list())
    elif subscriber_filter and subscriber_filter != 'all':
        query = query.filter(SubscriberName=subscriber_filter)

    return query.count()

# MODIFIED: Added subscriber_filter
def get_top_subscriber(start_date, end_date, subscriber_filter=None, user=None):
    """Get the top subscriber by usage count within the filtered group."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            return subscriber_filter # If filtering for one, they are the top one
        
        top_subscriber = query.values('SubscriberName').annotate(
            usage_count=Count('SearchIdentity')
        ).order_by('-usage_count').first()
        
        return top_subscriber['SubscriberName'] if top_subscriber else 'N/A'
    except Exception as e:
        logger.error(f"Error getting top subscriber: {str(e)}")
        return 'N/A'

# MODIFIED: Added subscriber_filter
def get_unique_products(subscriber_filter=None, user=None):
    """Get list of unique product names for filter dropdown, optionally filtered."""
    try:
        query = Usagereport.objects
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)
            
        products = query.values_list('ProductName', flat=True).distinct().order_by('ProductName')
        # return [product for product in products if product]
        return list(products)
    except Exception as e:
        logger.error(f"Error getting unique products: {str(e)}")
        return []

def get_unique_subscribers(user=None):
    """Get list of ALL unique subscriber names for the global filter dropdown."""
    try:
        query = Usagereport.objects
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
        subscribers = query.values_list('SubscriberName', flat=True).distinct().order_by('SubscriberName')
        return [subscriber for subscriber in subscribers if subscriber]
    except Exception as e:
        logger.error(f"Error getting unique subscribers: {str(e)}")
        return []

# MODIFIED: Simplified to one function and corrected logic
def get_top_subscribers_by_usage_filtered(start_date, end_date, subscriber_filter=None, limit=10, user=None):
    """Get top subscribers by usage volume with consistent filtering."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
        
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)
        
        top_subscribers = query.values('SubscriberName').annotate(
            usage_count=Count('SearchIdentity')
        ).order_by('-usage_count')
        
        if limit:
            top_subscribers = top_subscribers[:limit]
            
        return list(top_subscribers)
    except Exception as e:
        logger.error(f"Error getting filtered subscribers: {str(e)}")
        return []

# In views.py
def get_all_products_by_frequency(start_date, end_date, subscriber_filter=None, user=None):
    """Gets ALL products by frequency with consistent filtering, without a limit."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)

        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)

        all_products = query.values('ProductName').annotate(
            frequency=Count('ProductName')
        ).order_by('-frequency')

        return list(all_products)
    except Exception as e:
        logger.error(f"Error getting all products by frequency: {str(e)}")
        return []

# MODIFIED: Simplified to one function and corrected logic
def get_top_products_by_frequency_filtered(start_date, end_date, product_filter=None, subscriber_filter=None, limit=25, user=None):
    """Get top products by frequency with consistent filtering."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
        
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)

        if product_filter and product_filter != 'all':
            query = query.filter(ProductName__icontains=product_filter)
        
        top_products = query.values('ProductName').annotate(
            frequency=Count('ProductName')
        ).order_by('-frequency')[:limit]
        
        return list(top_products)
    except Exception as e:
        logger.error(f"Error getting filtered top products: {str(e)}")
        return []

# MODIFIED: Added subscriber_filter and corrected period logic
def get_churn_data_filtered(start_date, end_date, churn_days=None, subscriber_filter=None, user=None):
    """Get churn data with optional day filtering - HIGHLY OPTIMIZED"""
    try:
        # Date period logic
        duration_days = (end_date - start_date).days
        previous_end = start_date - timedelta(days=1)
        previous_start = previous_end - timedelta(days=duration_days)

        # Base queries
        previous_period_query = Usagereport.objects.filter(DetailsViewedDate__range=[previous_start, previous_end])
        current_period_query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            previous_period_query = previous_period_query.filter(SubscriberName__in=assigned)
            current_period_query = current_period_query.filter(SubscriberName__in=assigned)

        # Apply subscriber filter if relevant (for key subscribers)
        if subscriber_filter == 'key_subscribers':
            previous_period_query = previous_period_query.filter(SubscriberName__in=get_key_subscribers_list())
            current_period_query = current_period_query.filter(SubscriberName__in=get_key_subscribers_list())
        # Note: Churn for a single subscriber is not meaningful, so we don't handle that case.

        previous_subscribers = set(previous_period_query.values_list('SubscriberName', flat=True).distinct())
        current_subscribers = set(current_period_query.values_list('SubscriberName', flat=True).distinct())

        churned_subscribers_set = previous_subscribers - current_subscribers
        
        churned_count = len(churned_subscribers_set)
        previous_subscribers_count = len(previous_subscribers)
        current_subscribers_count = len(current_subscribers)
        
        churn_rate = (churned_count / previous_subscribers_count * 100) if previous_subscribers_count else 0
        
        return {
            'churned_count': churned_count,
            'churn_rate': round(churn_rate, 2),
            'previous_subscribers': previous_subscribers_count,
            'current_subscribers': current_subscribers_count
        }
    except Exception as e:
        logger.error(f"Error getting filtered churn data: {str(e)}")
        return {'churned_count': 0, 'churn_rate': 0, 'previous_subscribers': 0, 'current_subscribers': 0}

# MODIFIED: Added subscriber_filter
def get_retention_rate(start_date, end_date, subscriber_filter=None, user=None):
    """Calculate subscriber retention rate with filtering."""
    try:
        duration_days = (end_date - start_date).days
        previous_end = start_date - timedelta(days=1)
        previous_start = previous_end - timedelta(days=duration_days)

        previous_period_query = Usagereport.objects.filter(DetailsViewedDate__range=[previous_start, previous_end])
        current_period_query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            previous_period_query = previous_period_query.filter(SubscriberName__in=assigned)
            current_period_query = current_period_query.filter(SubscriberName__in=assigned)

        if subscriber_filter == 'key_subscribers':
            previous_period_query = previous_period_query.filter(SubscriberName__in=get_key_subscribers_list())
            current_period_query = current_period_query.filter(SubscriberName__in=get_key_subscribers_list())
        
        previous_subscribers = set(previous_period_query.values_list('SubscriberName', flat=True).distinct())
        current_subscribers = set(current_period_query.values_list('SubscriberName', flat=True).distinct())
        
        retained_subscribers = previous_subscribers.intersection(current_subscribers)
        retention_rate = len(retained_subscribers) / len(previous_subscribers) * 100 if previous_subscribers else 0
        
        return {
            'retention_rate': round(retention_rate, 2),
            'retained_count': len(retained_subscribers),
            'previous_count': len(previous_subscribers)
        }
    except Exception as e:
        logger.error(f"Error getting retention rate: {e}")
        return {'retention_rate': 0}


def get_engagement_rate(start_date, end_date, subscriber_filter=None, user=None):
    """Calculate engagement rate: Active subscribers / Total subscribers × 100.
    
    Shows what percentage of the total subscriber portfolio is actively using
    the service in the selected period.
    """
    try:
        total = get_total_subscribers(subscriber_filter, user)
        active = get_active_subscribers(start_date, end_date, subscriber_filter, user)
        
        engagement_rate = (active / total * 100) if total > 0 else 0
        
        return {
            'engagement_rate': round(engagement_rate, 2),
            'active_count': active,
            'total_count': total
        }
    except Exception as e:
        logger.error(f"Error getting engagement rate: {e}")
        return {'engagement_rate': 0, 'active_count': 0, 'total_count': 0}


# MODIFIED: This function is now correct
def get_usage_trends_filtered(start_date, end_date, usage_trends_days=None, subscriber_filter=None, user=None):
    """Get usage trends with optional day filtering and subscriber filtering - Returns COMPLETE date range including 0s"""
    try:
        # 1. Determine start date
        if usage_trends_days and usage_trends_days.isdigit():
            days = int(usage_trends_days)
            trends_start = end_date - timedelta(days=days)
        else:
            trends_start = start_date
            
        # 2. Base Query
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[trends_start, end_date]
        )
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
        
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)
        
        # 3. Get generic data (sparse)
        sparse_data = list(query.annotate(
            date=Cast('DetailsViewedDate', DateField())
        ).values('date').annotate(
            count=Count('ProductName')
        ).order_by('date'))
        
        # 4. Convert to dictionary for easy lookup
        # Keys are datetime.date objects from the query
        data_map = {item['date']: item['count'] for item in sparse_data}
        
        # 5. Build complete list with 0s for missing days
        complete_data = []
        current_date = trends_start
        
        # Ensure we are working with date objects for the loop
        if isinstance(current_date, datetime):
            current_date = current_date.date()
        if isinstance(end_date, datetime):
            end_date = end_date.date()

        while current_date <= end_date:
            # Use the date object directly for lookup
            count = data_map.get(current_date, 0)
            
            complete_data.append({
                'date': current_date.strftime('%Y-%m-%d'), # Return string for consistent JSON serialization
                'count': count
            })
            current_date += timedelta(days=1)
            
        return complete_data

    except Exception as e:
        logger.error(f"Error getting filtered usage trends: {str(e)}")
        return []


def get_all_subscribers_by_usage(start_date, end_date, subscriber_filter=None, user=None):
    """Gets ALL subscribers by usage with consistent filtering, without a limit."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)

        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)

        all_subscribers = query.values('SubscriberName').annotate(
            usage_count=Count('SubscriberName')  # Counting SubscriberName to include all records
        ).order_by('-usage_count')

        return list(all_subscribers)
    except Exception as e:
        logger.error(f"Error getting all subscribers by usage: {str(e)}")
        return []
def get_new_subscribers_trend_filtered(start_date, end_date, new_subscribers_days=None, subscriber_filter=None, user=None):
    """Get new subscribers trend with optional filtering."""
    try:
        if new_subscribers_days and new_subscribers_days.isdigit():
            days = int(new_subscribers_days)
            trends_start = end_date - timedelta(days=days)
        else:
            trends_start = start_date
        
        # Base query to find the first usage date for each subscriber
        first_usage_query = Usagereport.objects.values('SubscriberName').annotate(first_usage=Min('DetailsViewedDate'))
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            first_usage_query = first_usage_query.filter(SubscriberName__in=assigned)

        # Apply subscriber filter if provided
        if subscriber_filter == 'key_subscribers':
            first_usage_query = first_usage_query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
             first_usage_query = first_usage_query.filter(SubscriberName=subscriber_filter)

        # Filter the results to the desired date range
        first_usage_dates = first_usage_query.filter(first_usage__range=[trends_start, end_date])
        
        # Group by date in memory
        trend_data = defaultdict(int)
        for item in first_usage_dates:
            first_usage_date = item['first_usage']
            if isinstance(first_usage_date, datetime):
                first_usage_date = first_usage_date.date()
            if isinstance(first_usage_date, date):
                trend_data[first_usage_date] += 1
        
        # Create complete date range with zero counts for missing dates
        new_subscribers_by_date = []
        current_date = trends_start
        while current_date <= end_date:
            new_subscribers_by_date.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'new_subscribers': trend_data.get(current_date, 0)
            })
            current_date += timedelta(days=1)
        
        return new_subscribers_by_date
    except Exception as e:
        logger.error(f"Error getting filtered new subscribers trend: {str(e)}")
        return []

# MODIFIED: Added subscriber_filter
def get_highest_product_by_transaction(start_date, end_date, subscriber_filter=None, user=None):
    """Get the product with the highest number of transactions within the filtered group."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        if user and not user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
            query = query.filter(SubscriberName__in=assigned)
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=get_key_subscribers_list())
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)

        # In views.py
        top_product = query.values('ProductName').annotate(
            transaction_count=Count('SearchIdentity') # Corrected field
        ).order_by('-transaction_count').first()
        
        return top_product['ProductName'] if top_product else 'N/A'
    except Exception as e:
        logger.error(f"Error getting highest product by transaction: {str(e)}")
        return 'N/A'

# MODIFIED: Added subscriber_filter and corrected logic
def get_three_month_rolling_usage(subscriber_filter=None):
    """Get usage data for the current month and two previous months, broken down by month."""
    try:
        today = timezone.now().date()
        months_data = []
        
        for i in range(3): # Loop for 3 months
            # Determine the start and end of the month
            first_day_of_month = (today.replace(day=1) - timedelta(days=i*30)).replace(day=1)
            last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

            query = Usagereport.objects.filter(DetailsViewedDate__range=[first_day_of_month, last_day_of_month])
            
            if subscriber_filter == 'key_subscribers':
                query = query.filter(SubscriberName__in=get_key_subscribers_list())
            elif subscriber_filter and subscriber_filter != 'all':
                query = query.filter(SubscriberName=subscriber_filter)
            
            usage_count = query.count()
            month_name = first_day_of_month.strftime("%B")
            
            months_data.append({
                'month': f"{month_name} {first_day_of_month.year}",
                'usage_count': usage_count,
                'month_short': month_name[:3],
                'year': first_day_of_month.year
            })
        
        return list(reversed(months_data)) # Show oldest to newest
        
    except Exception as e:
        logger.error(f"Error getting 3-month rolling usage: {str(e)}")
        return []

# MODIFIED: Corrected filter logic
def get_daily_comparison(subscriber_filter=None, user=None):
    """Returns usage counts for yesterday and the same day of the previous month."""
    yesterday = (timezone.now() - timedelta(days=1)).date()
    # Safely get the same day last month
    try:
        prev_month_day = yesterday.replace(month=yesterday.month - 1)
    except ValueError: # Handles month-end cases like March 31 -> February
        prev_month_day = yesterday - timedelta(days=28)
        prev_month_day = prev_month_day.replace(day=yesterday.day)

    query_yesterday = Usagereport.objects.filter(DetailsViewedDate=yesterday)
    query_prev = Usagereport.objects.filter(DetailsViewedDate=prev_month_day)
    if user and not user.is_superuser:
        assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
        query_yesterday = query_yesterday.filter(SubscriberName__in=assigned)
        query_prev = query_prev.filter(SubscriberName__in=assigned)

    if subscriber_filter == 'key_subscribers':
        key_subs = get_key_subscribers_list()
        query_yesterday = query_yesterday.filter(SubscriberName__in=key_subs)
        query_prev = query_prev.filter(SubscriberName__in=key_subs)
    elif subscriber_filter and subscriber_filter != 'all':
        query_yesterday = query_yesterday.filter(SubscriberName=subscriber_filter)
        query_prev = query_prev.filter(SubscriberName=subscriber_filter)

    return {
        'yesterday': {
            'date': yesterday.strftime('%Y-%m-%d'),
            'count': query_yesterday.count(),
        },
        'previous_month_same_day': {
            'date': prev_month_day.strftime('%Y-%m-%d'),
            'count': query_prev.count(),
        }
    }


def get_week_comparison(subscriber_filter=None, user=None):
    """Returns usage counts for current week vs previous week with percentage change."""
    today = timezone.now().date()
    
    # Current week: Monday to today
    current_week_start = today - timedelta(days=today.weekday())
    current_week_end = today
    
    # Previous week: Last Monday to Sunday
    prev_week_start = current_week_start - timedelta(days=7)
    prev_week_end = current_week_start - timedelta(days=1)
    
    query_current = Usagereport.objects.filter(
        DetailsViewedDate__gte=current_week_start,
        DetailsViewedDate__lte=current_week_end
    )
    query_prev = Usagereport.objects.filter(
        DetailsViewedDate__gte=prev_week_start,
        DetailsViewedDate__lte=prev_week_end
    )
    
    if user and not user.is_superuser:
        assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
        query_current = query_current.filter(SubscriberName__in=assigned)
        query_prev = query_prev.filter(SubscriberName__in=assigned)
    
    if subscriber_filter == 'key_subscribers':
        key_subs = get_key_subscribers_list()
        query_current = query_current.filter(SubscriberName__in=key_subs)
        query_prev = query_prev.filter(SubscriberName__in=key_subs)
    elif subscriber_filter and subscriber_filter != 'all':
        query_current = query_current.filter(SubscriberName=subscriber_filter)
        query_prev = query_prev.filter(SubscriberName=subscriber_filter)
    
    current_count = query_current.count()
    prev_count = query_prev.count()
    
    # Calculate percentage change
    if prev_count > 0:
        percentage_change = round(((current_count - prev_count) / prev_count) * 100, 1)
    else:
        percentage_change = 100.0 if current_count > 0 else 0.0
    
    return {
        'current_week': {
            'start_date': current_week_start.strftime('%Y-%m-%d'),
            'end_date': current_week_end.strftime('%Y-%m-%d'),
            'count': current_count,
        },
        'previous_week': {
            'start_date': prev_week_start.strftime('%Y-%m-%d'),
            'end_date': prev_week_end.strftime('%Y-%m-%d'),
            'count': prev_count,
        },
        'percentage_change': percentage_change,
        'trend': 'up' if percentage_change > 0 else ('down' if percentage_change < 0 else 'same')
    }


def get_month_comparison(subscriber_filter=None, user=None):
    """Returns usage counts for current month vs previous month with percentage change."""
    today = timezone.now().date()
    
    # Current month: 1st to today
    current_month_start = today.replace(day=1)
    current_month_end = today
    
    # Previous month
    prev_month_end = current_month_start - timedelta(days=1)
    prev_month_start = prev_month_end.replace(day=1)
    
    query_current = Usagereport.objects.filter(
        DetailsViewedDate__gte=current_month_start,
        DetailsViewedDate__lte=current_month_end
    )
    query_prev = Usagereport.objects.filter(
        DetailsViewedDate__gte=prev_month_start,
        DetailsViewedDate__lte=prev_month_end
    )
    
    if user and not user.is_superuser:
        assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
        query_current = query_current.filter(SubscriberName__in=assigned)
        query_prev = query_prev.filter(SubscriberName__in=assigned)
    
    if subscriber_filter == 'key_subscribers':
        key_subs = get_key_subscribers_list()
        query_current = query_current.filter(SubscriberName__in=key_subs)
        query_prev = query_prev.filter(SubscriberName__in=key_subs)
    elif subscriber_filter and subscriber_filter != 'all':
        query_current = query_current.filter(SubscriberName=subscriber_filter)
        query_prev = query_prev.filter(SubscriberName=subscriber_filter)
    
    current_count = query_current.count()
    prev_count = query_prev.count()
    
    # Calculate percentage change
    if prev_count > 0:
        percentage_change = round(((current_count - prev_count) / prev_count) * 100, 1)
    else:
        percentage_change = 100.0 if current_count > 0 else 0.0
    
    return {
        'current_month': {
            'start_date': current_month_start.strftime('%Y-%m-%d'),
            'end_date': current_month_end.strftime('%Y-%m-%d'),
            'count': current_count,
            'month_name': current_month_start.strftime('%B %Y'),
        },
        'previous_month': {
            'start_date': prev_month_start.strftime('%Y-%m-%d'),
            'end_date': prev_month_end.strftime('%Y-%m-%d'),
            'count': prev_count,
            'month_name': prev_month_start.strftime('%B %Y'),
        },
        'percentage_change': percentage_change,
        'trend': 'up' if percentage_change > 0 else ('down' if percentage_change < 0 else 'same')
    }


# COMMENTED OUT - Not needed for now
# def get_same_day_comparison(subscriber_filter=None, user=None):
#     """Returns usage counts for today vs same day last month with percentage change."""
#     today = timezone.now().date()
#     
#     # Get same day last month
#     try:
#         if today.month == 1:
#             same_day_last_month = today.replace(year=today.year - 1, month=12)
#         else:
#             same_day_last_month = today.replace(month=today.month - 1)
#     except ValueError:
#         # Handle month-end edge cases (e.g., March 31 -> Feb doesn't have 31)
#         # Go back one month and use the last day of that month
#         if today.month == 1:
#             prev_month = today.replace(year=today.year - 1, month=12, day=1)
#         else:
#             prev_month = today.replace(month=today.month - 1, day=1)
#         # Get last day of that month
#         import calendar
#         last_day = calendar.monthrange(prev_month.year, prev_month.month)[1]
#         same_day_last_month = prev_month.replace(day=min(today.day, last_day))
#     
#     query_today = Usagereport.objects.filter(DetailsViewedDate=today)
#     query_prev = Usagereport.objects.filter(DetailsViewedDate=same_day_last_month)
#     
#     if user and not user.is_superuser:
#         assigned = list(Subscriber.objects.filter(managed_by=user).values_list('name', flat=True))
#         query_today = query_today.filter(SubscriberName__in=assigned)
#         query_prev = query_prev.filter(SubscriberName__in=assigned)
#     
#     if subscriber_filter == 'key_subscribers':
#         key_subs = get_key_subscribers_list()
#         query_today = query_today.filter(SubscriberName__in=key_subs)
#         query_prev = query_prev.filter(SubscriberName__in=key_subs)
#     elif subscriber_filter and subscriber_filter != 'all':
#         query_today = query_today.filter(SubscriberName=subscriber_filter)
#         query_prev = query_prev.filter(SubscriberName=subscriber_filter)
#     
#     today_count = query_today.count()
#     prev_count = query_prev.count()
#     
#     # Calculate percentage change
#     if prev_count > 0:
#         percentage_change = round(((today_count - prev_count) / prev_count) * 100, 1)
#     else:
#         percentage_change = 100.0 if today_count > 0 else 0.0
#     
#     return {
#         'today': {
#             'date': today.strftime('%Y-%m-%d'),
#             'count': today_count,
#         },
#         'same_day_last_month': {
#             'date': same_day_last_month.strftime('%Y-%m-%d'),
#             'count': prev_count,
#         },
#         'percentage_change': percentage_change,
#         'trend': 'up' if percentage_change > 0 else ('down' if percentage_change < 0 else 'same')
#     }


# --- The remaining functions like download_churned_subscribers etc. are not part of the dashboard API flow ---
# --- and do not need modification for this specific task. They remain as they are. ---

def get_churned_subscribers_list(start_date, end_date, churn_days=None):
    """Get list of churned subscriber names for download"""
    try:
        # Use custom days if provided, otherwise use default logic
        if churn_days and churn_days.isdigit():
            days = int(churn_days)
            analysis_start = end_date - timedelta(days=days)
            previous_start = analysis_start - timedelta(days=days)
            previous_end = analysis_start
        else:
            # Use original logic
            analysis_start = start_date
            previous_start = start_date - timedelta(days=(end_date - start_date).days)
            previous_end = start_date
        
        # Get previous subscribers
        previous_subscribers = set(Usagereport.objects.filter(
            DetailsViewedDate__range=[previous_start, previous_end]
        ).values_list('SubscriberName', flat=True).distinct())
        
        # Get current subscribers
        current_subscribers = set(Usagereport.objects.filter(
            DetailsViewedDate__range=[analysis_start, end_date]
        ).values_list('SubscriberName', flat=True).distinct())
        
        # Calculate churned subscribers
        churned_subscribers = previous_subscribers - current_subscribers
        
        # Return sorted list
        return sorted(list(churned_subscribers))
        
    except Exception as e:
        logger.error(f"Error getting churned subscribers list: {str(e)}")
        return []

# In views.py, add this new function at the end of the file.

# @login_required
# In views.py, replace the previous download_top_subscribers_csv function

# @login_required
# def download_top_subscribers_csv(request):
#     """
#     Generates and serves a TSV (Tab-Separated Values) file with a detailed breakdown of usage 
#     for the top 10 subscribers, respecting all active dashboard filters.
#     """
#     try:
#         # Get and validate filters from the request
#         start_date_str = request.GET.get('start_date')
#         end_date_str = request.GET.get('end_date')
#         subscriber_filter = request.GET.get('subscriber_filter')

#         today = timezone.now().date()
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
        
#         if end_date_str:
#             end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
#         else:
#             end_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

#         # Build the base queryset with date and subscriber filters
#         base_query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])

#         if subscriber_filter == 'key_subscribers':
#             base_query = base_query.filter(SubscriberName__in=get_key_subscribers_list())
#         elif subscriber_filter and subscriber_filter != 'all':
#             base_query = base_query.filter(SubscriberName=subscriber_filter)

#         # First, find the names of the top 10 subscribers within the filtered group
#         top_subscribers_names = list(base_query.values('SubscriberName')
#                                     .annotate(total_usage=Count('SearchIdentity'))
#                                     .order_by('-total_usage')[:10]
#                                     .values_list('SubscriberName', flat=True))

#         if not top_subscribers_names:
#             return HttpResponse("No data available for the selected filters.", content_type="text/plain")

#         # Now, get the detailed product breakdown for ONLY those top subscribers
#         detailed_data = base_query.filter(SubscriberName__in=top_subscribers_names) \
#                                   .values('SubscriberName', 'ProductName') \
#                                   .annotate(usage_count=Count('SearchIdentity')) \
#                                   .order_by('SubscriberName', '-usage_count')

#         # Generate the TSV (tab-separated) response
#         response = HttpResponse(content_type='text/tab-separated-values')
#         filename = f"top_subscribers_usage_detail_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.tsv"
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'

#         writer = csv.writer(response, delimiter='\t')
#         writer.writerow(['Subscriber Name', 'Product Name', 'Usage Count'])

#         for row in detailed_data:
#             writer.writerow([
#                 row['SubscriberName'],
#                 row['ProductName'],
#                 row['usage_count']
#             ])

#         return response

#     except Exception as e:
#         logger.error(f"Error downloading top subscribers TSV: {e}")
#         return HttpResponse("An error occurred while generating the report.", status=500)
# In views.py, replace the entire download_top_subscribers_csv function.

# @login_required
# def download_top_subscribers_csv(request):
#     """
#     Generates a TSV file for top subscriber usage.
#     - If a specific subscriber is selected, it downloads their detailed usage.
#     - Otherwise, it downloads the detailed usage for the top 10 subscribers in the current filter group.
#     """
#     try:
#         # Get filters from the request
#         start_date_str = request.GET.get('start_date')
#         end_date_str = request.GET.get('end_date')
#         global_subscriber_filter = request.GET.get('subscriber_filter')
#         selected_subscriber = request.GET.get('selected_subscriber', None)

#         # Handle date ranges
#         today = timezone.now().date()
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
#         if end_date_str:
#             end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
#         else:
#             end_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

#         # Build the base queryset with the date filter
#         base_query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])
#         filename_part = ""

#         if selected_subscriber:
#             # Case 1: A specific subscriber was selected from the dropdown.
#             # Filter for that single subscriber's detailed product usage.
#             detailed_data = base_query.filter(SubscriberName=selected_subscriber) \
#                                       .values('SubscriberName', 'ProductName') \
#                                       .annotate(usage_count=Count('SearchIdentity')) \
#                                       .order_by('-usage_count')
#             filename_part = clean_filename(selected_subscriber)
#         else:
#             # Case 2: No specific subscriber selected, so get the top 10 from the global filter group.
            
#             # Apply the global filter to the base query
#             if global_subscriber_filter == 'key_subscribers':
#                 base_query = base_query.filter(SubscriberName__in=get_key_subscribers_list())
#                 filename_part = "Key_Subscribers_Top_10"
#             elif global_subscriber_filter and global_subscriber_filter != 'all':
#                 base_query = base_query.filter(SubscriberName=global_subscriber_filter)
#                 filename_part = clean_filename(global_subscriber_filter)
#             else:
#                 filename_part = "All_Subscribers_Top_10"

#             # Get the names of the top 10 subscribers from the filtered group
#             top_subscribers_names = list(base_query.values('SubscriberName')
#                                         .annotate(total_usage=Count('SearchIdentity'))
#                                         .order_by('-total_usage')[:10]
#                                         .values_list('SubscriberName', flat=True))

#             if not top_subscribers_names:
#                 return HttpResponse("No data available for the selected filters.", content_type="text/plain")

#             # Get the detailed product usage for ONLY those top 10 subscribers
#             detailed_data = base_query.filter(SubscriberName__in=top_subscribers_names) \
#                                       .values('SubscriberName', 'ProductName') \
#                                       .annotate(usage_count=Count('SearchIdentity')) \
#                                       .order_by('SubscriberName', '-usage_count')
        
#         # Generate the TSV response
#         response = HttpResponse(content_type='text/tab-separated-values')
#         filename = f"{filename_part}_usage_detail_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.tsv"
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'

#         writer = csv.writer(response, delimiter='\t')
#         writer.writerow(['Subscriber Name', 'Product Name', 'Usage Count'])

#         for row in detailed_data:
#             writer.writerow([
#                 row['SubscriberName'],
#                 row['ProductName'],
#                 row['usage_count']
#             ])

#         return response

#     except Exception as e:
#         logger.error(f"Error downloading top subscribers TSV: {e}")
#         return HttpResponse("An error occurred while generating the report.", status=500)

# In views.py, replace the download_top_subscribers_csv function

@login_required
def download_top_subscribers_csv(request):
    """
    Generates a TSV file for top subscriber usage. This version is optimized to
    prevent database timeouts by ensuring querysets are constructed efficiently.
    """
    try:
        # 1. Get and parse request parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        global_subscriber_filter = request.GET.get('subscriber_filter')
        selected_subscriber = request.GET.get('selected_subscriber', None)

        today = timezone.now().date()
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        filename_part = ""
        
        # 2. Determine the query based on parameters
        if selected_subscriber:
            # Case 1: A specific subscriber was selected.
            filename_part = clean_filename(selected_subscriber)
            detailed_data = Usagereport.objects.filter(
                DetailsViewedDate__range=[start_date, end_date],
                SubscriberName=selected_subscriber
            ).values('SubscriberName', 'ProductName') \
             .annotate(usage_count=Count('SearchIdentity')) \
             .order_by('-usage_count')

        else:
            # Case 2: No specific subscriber, so find the Top 10 for the global filter group.
            query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])
            
            if global_subscriber_filter == 'key_subscribers':
                query = query.filter(SubscriberName__in=get_key_subscribers_list())
                filename_part = "Key_Subscribers_Top_10"
            elif global_subscriber_filter and global_subscriber_filter != 'all':
                query = query.filter(SubscriberName=global_subscriber_filter)
                filename_part = clean_filename(global_subscriber_filter)
            else:
                filename_part = "All_Subscribers_Top_10"

            # This is the potentially slow query. The indexes will speed it up.
            top_subscribers_names = list(query.values('SubscriberName')
                                        .annotate(total_usage=Count('SearchIdentity'))
                                        .order_by('-total_usage')[:10]
                                        .values_list('SubscriberName', flat=True))

            if not top_subscribers_names:
                return HttpResponse("No data available for the selected filters.", content_type="text/plain")

            detailed_data = query.filter(SubscriberName__in=top_subscribers_names) \
                                 .values('SubscriberName', 'ProductName') \
                                 .annotate(usage_count=Count('SearchIdentity')) \
                                 .order_by('SubscriberName', '-usage_count')

        # 3. Generate and stream the TSV response
        response = HttpResponse(content_type='text/tab-separated-values')
        filename = f"{filename_part}_usage_detail_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.tsv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response, delimiter='\t')
        writer.writerow(['Subscriber Name', 'Product Name', 'Usage Count'])

        for row in detailed_data:
            writer.writerow([row['SubscriberName'], row['ProductName'], row['usage_count']])

        return response

    except Exception as e:
        logger.error(f"Error downloading top subscribers TSV: {e}")
        return HttpResponse("An error occurred while generating the report.", status=500)


def download_churned_subscribers(request):
    """Download churned subscribers as CSV file"""

    
    try:
        # Get date range from request
        start_date_str = request.GET.get('start_date', None)
        end_date_str = request.GET.get('end_date', None)
        churn_days = request.GET.get('churn_days', None)
        
        # Parse dates or use defaults
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            # Default to first day of current month
            today = timezone.now().date()
            start_date = today.replace(day=1)
            
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            # Default to first day of next month
            today = timezone.now().date()
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1)
        
        # Get churned subscribers list
        churned_subscribers = get_churned_subscribers_list(start_date, end_date, churn_days)
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="churned_subscribers_{start_date}_{end_date}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Subscriber Name'])  # Header
        
        # Write subscriber names
        for subscriber in churned_subscribers:
            writer.writerow([subscriber])
        
        return response
        
    except Exception as e:
        logger.error(f"Error downloading churned subscribers: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)


def download_new_subscribers(request):
    """Download new subscribers as TXT file with subscriber names and join dates"""
    from django.http import HttpResponse
    from datetime import datetime
    
    try:
        # Get date range from request
        start_date_str = request.GET.get('start_date', None)
        end_date_str = request.GET.get('end_date', None)
        
        # Parse dates or use defaults
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            # Default to 30 days ago
            today = timezone.now().date()
            start_date = today - timedelta(days=30)
            
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            # Default to today
            end_date = timezone.now().date()
        
        # Get new subscribers details with names and join dates
        new_subscribers_details = get_new_subscribers_details(start_date, end_date)
        
        # Create TXT response
        response = HttpResponse(content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="new_subscribers_{start_date}_{end_date}.txt"'
        
        # Write header
        content = f"New Subscribers Report\n"
        content += f"Date Range: {start_date} to {end_date}\n"
        content += f"Total New Subscribers: {len(new_subscribers_details)}\n\n"
        content += f"{'Subscriber Name':<50} {'Date Joined':<15}\n"
        content += f"{'-' * 50} {'-' * 15}\n"
        
        # Write subscriber details
        for subscriber in new_subscribers_details:
            content += f"{subscriber['name']:<50} {subscriber['date_joined']:<15}\n"
        
        response.write(content)
        return response
        
    except Exception as e:
        logger.error(f"Error downloading new subscribers: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)


def new_subscribers_trend_api(request):
    """API endpoint for new subscribers trend data with custom date range"""
    try:
        # Get date range from request
        start_date_str = request.GET.get('start_date', None)
        end_date_str = request.GET.get('end_date', None)
        
        if not start_date_str or not end_date_str:
            return JsonResponse({'error': 'Both start_date and end_date are required'}, status=400)
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Validate date range
        if start_date > end_date:
            return JsonResponse({'error': 'Start date cannot be later than end date'}, status=400)
        
        # Get new subscribers data
        new_subscribers_data = get_new_subscribers_trend_optimized(start_date, end_date)
        
        return JsonResponse({
            'new_subscribers': new_subscribers_data,
            'start_date': start_date_str,
            'end_date': end_date_str
        })
        
    except ValueError as e:
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    except Exception as e:
        logger.error(f"Error in new subscribers trend API: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


def usage_trends_api(request):
    """API endpoint for usage trends data with flexible comparison modes.
    
    Comparison modes:
    - 'default': Current month vs Previous month (dashboard default)
    - 'preset': Current period vs Previous equivalent period (for 7/30/90 day filters)
    - 'custom': For custom date ranges - auto-detects if two months and splits, 
                or shows single line for single month ranges
    """
    try:
        # Get parameters from request
        start_date_str = request.GET.get('start_date', None)
        end_date_str = request.GET.get('end_date', None)
        comparison_mode = request.GET.get('comparison_mode', 'default')
        
        # Get subscriber filter if provided
        subscriber_filter = request.GET.get('subscriber_filter', 'all')
        if subscriber_filter == 'all':
            subscriber_filter = None
        
        today = timezone.now().date()
        
        # Handle different comparison modes
        if comparison_mode == 'default':
            # Default: Current month vs Previous month
            current_month_start = today.replace(day=1)
            current_month_end = today
            
            # Previous month
            prev_month_end = current_month_start - timedelta(days=1)
            prev_month_start = prev_month_end.replace(day=1)
            
            # Get data for both months
            current_data = get_usage_trends_filtered(current_month_start, current_month_end, None, subscriber_filter)
            previous_data = get_usage_trends_filtered(prev_month_start, prev_month_end, None, subscriber_filter)
            
            # Format labels
            current_label = f"{current_month_start.strftime('%B %Y')}"
            previous_label = f"{prev_month_start.strftime('%B %Y')}"
            
            return JsonResponse({
                'usage_trends': current_data,
                'usage_trends_previous': previous_data,
                'current_period_label': current_label,
                'previous_period_label': previous_label,
                'start_date': current_month_start.strftime('%Y-%m-%d'),
                'end_date': current_month_end.strftime('%Y-%m-%d'),
                'has_comparison': True
            })
        
        elif comparison_mode == 'preset':
            # Preset filters: Current period vs Previous equivalent period
            if not start_date_str or not end_date_str:
                return JsonResponse({'error': 'Both start_date and end_date are required for preset mode'}, status=400)
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            if start_date > end_date:
                return JsonResponse({'error': 'Start date cannot be later than end date'}, status=400)
            
            # Calculate period duration and previous period
            period_duration = (end_date - start_date).days + 1
            prev_end_date = start_date - timedelta(days=1)
            prev_start_date = prev_end_date - timedelta(days=period_duration - 1)
            
            # Get data for both periods
            current_data = get_usage_trends_filtered(start_date, end_date, None, subscriber_filter)
            previous_data = get_usage_trends_filtered(prev_start_date, prev_end_date, None, subscriber_filter)
            
            # Format labels
            current_label = f"{start_date.strftime('%b %d')} - {end_date.strftime('%b %d, %Y')}"
            previous_label = f"{prev_start_date.strftime('%b %d')} - {prev_end_date.strftime('%b %d, %Y')}"
            
            return JsonResponse({
                'usage_trends': current_data,
                'usage_trends_previous': previous_data,
                'current_period_label': current_label,
                'previous_period_label': previous_label,
                'start_date': start_date_str,
                'end_date': end_date_str,
                'has_comparison': True
            })
        
        elif comparison_mode == 'custom':
            # Custom date range: Compare two FULL months when range spans two different months
            if not start_date_str or not end_date_str:
                return JsonResponse({'error': 'Both start_date and end_date are required for custom mode'}, status=400)
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            if start_date > end_date:
                return JsonResponse({'error': 'Start date cannot be later than end date'}, status=400)
            
            # Check if range spans two different calendar months
            start_month = (start_date.year, start_date.month)
            end_month = (end_date.year, end_date.month)
            
            if start_month != end_month:
                # Two-month comparison: Get FULL months for both
                # First month: Use the month of start_date (full month)
                import calendar
                first_month_start = start_date.replace(day=1)
                first_month_last_day = calendar.monthrange(start_date.year, start_date.month)[1]
                first_month_end = start_date.replace(day=first_month_last_day)
                
                # Second month: Use the month of end_date (full month)
                second_month_start = end_date.replace(day=1)
                second_month_last_day = calendar.monthrange(end_date.year, end_date.month)[1]
                second_month_end = end_date.replace(day=second_month_last_day)
                
                # Get data for FULL first month (previous/orange)
                first_month_data = get_usage_trends_filtered(first_month_start, first_month_end, None, subscriber_filter)
                # Get data for FULL second month (current/blue)
                second_month_data = get_usage_trends_filtered(second_month_start, second_month_end, None, subscriber_filter)
                
                # Format labels
                first_month_label = f"{first_month_start.strftime('%B %Y')}"
                second_month_label = f"{second_month_start.strftime('%B %Y')}"
                
                # DEBUG logging
                logger.info(f"[DEBUG] Custom 2-month mode: first={first_month_start} to {first_month_end} ({len(first_month_data)} days), second={second_month_start} to {second_month_end} ({len(second_month_data)} days)")
                print(f"[DEBUG] Custom 2-month mode: first={first_month_start} to {first_month_end} ({len(first_month_data)} days), second={second_month_start} to {second_month_end} ({len(second_month_data)} days)")
                
                return JsonResponse({
                    'usage_trends': second_month_data,  # Current = second/later month
                    'usage_trends_previous': first_month_data,  # Previous = first/earlier month
                    'current_period_label': second_month_label,
                    'previous_period_label': first_month_label,
                    'start_date': start_date_str,
                    'end_date': end_date_str,
                    'has_comparison': True
                })
            else:
                # Single month range: Compare against same days in previous month
                current_data = get_usage_trends_filtered(start_date, end_date, None, subscriber_filter)
                
                # Calculate previous month's equivalent dates
                first_of_current = start_date.replace(day=1)
                last_of_previous = first_of_current - timedelta(days=1)
                first_of_previous = last_of_previous.replace(day=1)
                
                # Match the same date range in the previous month
                # e.g. March 1-30 -> February 1-(28 or 29)
                prev_start_day = min(start_date.day, last_of_previous.day)
                prev_end_day = min(end_date.day, last_of_previous.day)
                
                prev_start_date = first_of_previous.replace(day=prev_start_day)
                prev_end_date = first_of_previous.replace(day=prev_end_day)
                
                # Get previous month data
                previous_data = get_usage_trends_filtered(prev_start_date, prev_end_date, None, subscriber_filter)
                
                # Format labels
                current_label = f"{start_date.strftime('%B %Y')}"
                previous_label = f"{prev_start_date.strftime('%B %Y')}"
                
                # DEBUG logging
                logger.info(f"[DEBUG] Custom single-month mode: current={start_date} to {end_date} ({len(current_data)} days), previous={prev_start_date} to {prev_end_date} ({len(previous_data)} days)")
                print(f"[DEBUG] Custom single-month mode: current={start_date} to {end_date} ({len(current_data)} days), previous={prev_start_date} to {prev_end_date} ({len(previous_data)} days)")
                
                return JsonResponse({
                    'usage_trends': current_data,
                    'usage_trends_previous': previous_data,
                    'current_period_label': current_label,
                    'previous_period_label': previous_label,
                    'start_date': start_date_str,
                    'end_date': end_date_str,
                    'has_comparison': True
                })

        
        else:
            return JsonResponse({'error': f'Invalid comparison_mode: {comparison_mode}'}, status=400)
        
    except ValueError as e:
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    except Exception as e:
        logger.error(f"Error in usage trends API: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)



def get_subscriber_product_rate(subscriber, product_key):
    """
    Get the subscriber-specific rate for a product with fallback to ENQUIRY_RATES.
    
    Args:
        subscriber (str): The subscriber name
        product_key (str): The product key in ENQUIRY_RATES format (with underscores)
    
    Returns:
        Decimal: The rate for the subscriber-product combination
    """
    try:
        # Map product keys to actual product names in database
        product_variations = [
            product_key.replace('_', ' '),
            product_key.replace('_', '-'),
            product_key.title().replace('_', ' '),
        ]
        
        # Try to find a subscriber-specific rate
        for variation in product_variations:
            try:
                # Use filter().first() to handle duplicate records gracefully
                subscriber_rate = SubscriberProductRate.objects.filter(
                    subscriber_name=subscriber,
                    product_name__icontains=variation
                ).first()
                if subscriber_rate:
                    return subscriber_rate.rate
            except Exception:
                continue
        
        # Fall back to ENQUIRY_RATES if no specific rate exists
        return ENQUIRY_RATES.get(product_key, Decimal('0.00'))
    except Exception as e:
        logger.error(f"Error getting subscriber product rate: {str(e)}")
        return ENQUIRY_RATES.get(product_key, Decimal('0.00'))

def get_all_subscriber_product_rate():
    """
    Gets all product subscriber rates in a single query and groups them in memory.
    """
    try:
        # Use iterator for memory efficiency on large datasets
        all_rates_data = SubscriberProductRate.objects.values_list(
            'subscriber_name', 'product_name', 'rate'
        ).iterator(chunk_size=2000)

        # Group rates by subscriber for fast lookups
        grouped_rates = defaultdict(dict)
        for subscriber_name, product_name, rate in all_rates_data:
            # Store as {subscriber: {product: rate}}
            grouped_rates[subscriber_name][product_name.lower()] = rate
            
        return grouped_rates
        
    except Exception as e:
        logger.error(f"Error getting all subscriber product rates: {str(e)}")
        return {} # Return an empty dictionary on error


def get_top_products_by_subscriber_filtered(start_date, end_date, subscriber_name):
    """Show only products used by selected subscriber"""
    try:
        from django.db.models import Count
        
        if not subscriber_name:
            return []
        
        # Get products used by the specific subscriber
        products = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date],
            SubscriberName=subscriber_name
        ).values('ProductName').annotate(
            usage_count=Count('ProductName')
        ).order_by('-usage_count')[:10]
        
        result = []
        for product in products:
            result.append({
                'product': product['ProductName'],
                'count': product['usage_count']
            })
        
        return result
    except Exception as e:
        # logger.error(f"Error getting top products by subscriber: {str(e)}")
        return []


def get_new_subscribers_details(start_date, end_date):
    """Get detailed list of new subscribers with names and join dates"""
    try:
        from datetime import datetime, date

        # Ensure start_date and end_date are consistently date objects
        if isinstance(start_date, datetime):
            start_date = start_date.date()
        elif isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

        if isinstance(end_date, datetime):
            end_date = end_date.date()
        elif isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Get first usage date for each subscriber with their names
        new_subscribers = Usagereport.objects.values('SubscriberName').annotate(
            first_usage=Min('DetailsViewedDate')
        ).filter(
            first_usage__range=[start_date, end_date]
        ).order_by('first_usage', 'SubscriberName')

        # Format the results
        subscribers_details = []
        for item in new_subscribers:
            first_usage = item['first_usage']
            if first_usage is None:
                continue

            # Ensure first_usage is a date object
            if isinstance(first_usage, datetime):
                first_usage = first_usage.date()
            elif isinstance(first_usage, str):
                first_usage = datetime.strptime(first_usage.split(' ')[0], '%Y-%m-%d').date()
            
            if isinstance(first_usage, date):
                subscribers_details.append({
                    'name': item['SubscriberName'],
                    'date_joined': first_usage.strftime('%Y-%m-%d')
                })
        
        return subscribers_details
    except Exception as e:
        logger.error(f"Error getting new subscribers details: {str(e)}")
        return []


def get_new_subscribers_trend_optimized(start_date, end_date):
    """Get new subscribers trend - OPTIMIZED version without day filtering"""
    try:
        

        # Ensure start_date and end_date are consistently date objects
        if isinstance(start_date, datetime):
            start_date = start_date.date()
        elif isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

        if isinstance(end_date, datetime):
            end_date = end_date.date()
        elif isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Single query to get first usage date for each subscriber
        first_usage_dates = Usagereport.objects.values('SubscriberName').annotate(
            first_usage=Min('DetailsViewedDate')
        ).filter(
            first_usage__range=[start_date, end_date]
        )

        # Group by date
        trend_data = {}
        for item in first_usage_dates:
            first_usage = item['first_usage']
            if first_usage is None:
                continue

            # Defensively ensure first_usage is a date object before calling strftime
            if isinstance(first_usage, datetime):
                first_usage = first_usage.date()
            elif isinstance(first_usage, str):
                # If the database returns a string, convert it. This handles formats like 'YYYY-MM-DD' or 'YYYY-MM-DD HH:MM:SS'
                first_usage = datetime.strptime(first_usage.split(' ')[0], '%Y-%m-%d').date()
            
            # Now we can safely call strftime
            if isinstance(first_usage, date):
                date_key = first_usage.strftime('%Y-%m-%d')
                trend_data[date_key] = trend_data.get(date_key, 0) + 1
        
        # Create complete date range with zero counts for missing dates
        new_subscribers_by_date = []
        current_date = start_date
        
        # The redundant checks inside the loop have been removed for cleanliness and efficiency.
        while current_date <= end_date:
            date_key = current_date.strftime('%Y-%m-%d')
            new_subscribers_by_date.append({
                'date': date_key,
                'new_subscribers': trend_data.get(date_key, 0)
            })
            current_date += timedelta(days=1)
        
        return new_subscribers_by_date
    except Exception as e:
        logger.error(f"Error getting new subscribers trend optimized: {str(e)}")
        return []


@login_required
def subscriber_performance(request):
    """Dedicated page for viewing subscriber performance metrics."""
    context = {
        'title': 'Subscriber Performance Analytics'
    }
    return render(request, 'bulkrep/subscriber_performance.html', context)


@login_required
def subscriber_performance_api(request):
    """API endpoint for subscriber performance with Period A vs Period B comparison."""
    try:
        import calendar
        
        # Get parameters
        mode = request.GET.get('mode', 'auto')  # auto or custom
        search_query = request.GET.get('search', '').strip()
        limit = int(request.GET.get('limit', 100))
        
        today = timezone.now().date()
        
        if mode == 'custom':
            # Custom Period A and Period B from user input
            period_a_start = request.GET.get('period_a_start')
            period_a_end = request.GET.get('period_a_end')
            period_b_start = request.GET.get('period_b_start')
            period_b_end = request.GET.get('period_b_end')
            
            if period_a_start and period_a_end and period_b_start and period_b_end:
                a_start = datetime.strptime(period_a_start, '%Y-%m-%d').date()
                a_end = datetime.strptime(period_a_end, '%Y-%m-%d').date()
                b_start = datetime.strptime(period_b_start, '%Y-%m-%d').date()
                b_end = datetime.strptime(period_b_end, '%Y-%m-%d').date()
                
                period_a_label = f"{a_start.strftime('%b %d')} - {a_end.strftime('%b %d, %Y')}"
                period_b_label = f"{b_start.strftime('%b %d')} - {b_end.strftime('%b %d, %Y')}"
            else:
                # Fallback to auto if dates missing
                mode = 'auto'
        
        if mode == 'auto':
            # Auto mode: Last full month (Period A) vs Current month to date (Period B)
            # Period B = Current month (1st to today)
            b_start = today.replace(day=1)
            b_end = today
            
            # Period A = Last full month
            a_end = b_start - timedelta(days=1)
            a_start = a_end.replace(day=1)
            
            period_a_label = f"{a_start.strftime('%B %Y')}"
            period_b_label = f"{b_start.strftime('%B %Y')} (to date)"
        
        # Build base query respecting user permissions
        base_query = Usagereport.objects
        if not request.user.is_superuser:
            assigned = list(Subscriber.objects.filter(managed_by=request.user).values_list('name', flat=True))
            base_query = base_query.filter(SubscriberName__in=assigned)
        
        # Apply search filter
        if search_query:
            base_query = base_query.filter(SubscriberName__icontains=search_query)
        
        # Get Period A counts
        period_a_counts = dict(
            base_query.filter(
                DetailsViewedDate__gte=a_start,
                DetailsViewedDate__lte=a_end
            ).values('SubscriberName').annotate(
                count=Count('SearchIdentity')
            ).values_list('SubscriberName', 'count')
        )
        
        # Get Period B counts
        period_b_counts = dict(
            base_query.filter(
                DetailsViewedDate__gte=b_start,
                DetailsViewedDate__lte=b_end
            ).values('SubscriberName').annotate(
                count=Count('SearchIdentity')
            ).values_list('SubscriberName', 'count')
        )
        
        # Combine all subscriber names
        all_subscribers = set(period_a_counts.keys()) | set(period_b_counts.keys())
        
        # Build performance data
        performance_data = []
        for subscriber_name in all_subscribers:
            a_count = period_a_counts.get(subscriber_name, 0)
            b_count = period_b_counts.get(subscriber_name, 0)
            
            # Calculate percentage change (B compared to A)
            if a_count > 0:
                percentage_change = round(((b_count - a_count) / a_count) * 100, 1)
            else:
                percentage_change = 100.0 if b_count > 0 else 0.0
            
            performance_data.append({
                'subscriber_name': subscriber_name,
                'period_a_usage': a_count,
                'period_b_usage': b_count,
                'percentage_change': percentage_change,
                'trend': 'up' if percentage_change > 0 else ('down' if percentage_change < 0 else 'same')
            })
        
        # Sort by Period B usage descending and limit
        performance_data.sort(key=lambda x: x['period_b_usage'], reverse=True)
        performance_data = performance_data[:limit]
        
        return JsonResponse({
            'success': True,
            'data': performance_data,
            'meta': {
                'mode': mode,
                'period_a_label': period_a_label,
                'period_b_label': period_b_label,
                'period_a_range': f"{a_start.strftime('%Y-%m-%d')} to {a_end.strftime('%Y-%m-%d')}",
                'period_b_range': f"{b_start.strftime('%Y-%m-%d')} to {b_end.strftime('%Y-%m-%d')}",
                'total_results': len(performance_data),
                'search_query': search_query,
            }
        })
        
    except Exception as e:
        logger.error(f"Error in subscriber_performance_api: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

