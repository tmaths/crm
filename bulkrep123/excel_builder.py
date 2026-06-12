"""
Excel report builder that reconstructs workbooks from template_structure.json.
Optimized for low memory usage and fast generation.
"""
import json
import os
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from django.conf import settings
import logging

logger = logging.getLogger(__name__)

# Global cache for template structure
_template_structure = None
_style_cache = {}


def load_template_structure():
    """Load and cache the template structure JSON."""
    global _template_structure
    
    if _template_structure is None:
        template_json_path = os.path.join(settings.BASE_DIR.parent, 'template_structure.json')
        try:
            with open(template_json_path, 'r', encoding='utf-8') as f:
                _template_structure = json.load(f)
            logger.info(f"Loaded template structure with {len(_template_structure.get('cells', []))} cells")
        except Exception as e:
            logger.error(f"Failed to load template structure: {e}")
            raise
    
    return _template_structure


def get_cached_style(font_dict, alignment_dict, fill_dict, border_dict, number_format):
    """
    Get or create a cached style object to avoid recreating identical styles.
    Returns a tuple of (font, alignment, fill, border, number_format).
    """
    # Create a hashable key from the style properties
    style_key = (
        tuple(sorted(font_dict.items())) if font_dict else None,
        tuple(sorted(alignment_dict.items())) if alignment_dict else None,
        tuple(sorted(fill_dict.items())) if fill_dict else None,
        tuple(sorted(border_dict.items())) if border_dict else None,
        number_format
    )
    
    if style_key in _style_cache:
        return _style_cache[style_key]
    
    # Create new style objects
    font = None
    if font_dict:
        font = Font(
            name=font_dict.get('name', 'Calibri'),
            size=font_dict.get('size', 11),
            bold=font_dict.get('bold', False),
            italic=font_dict.get('italic', False),
            color=font_dict.get('color', '00000000')
        )
    
    alignment = None
    if alignment_dict:
        alignment = Alignment(
            horizontal=alignment_dict.get('horizontal', 'general'),
            vertical=alignment_dict.get('vertical', 'bottom'),
            wrap_text=alignment_dict.get('wrap_text', False)
        )
    
    fill = None
    if fill_dict and fill_dict.get('pattern'):
        fill = PatternFill(
            fill_type=fill_dict.get('pattern', 'solid'),
            fgColor=fill_dict.get('fg_color', 'FFFFFFFF'),
            bgColor=fill_dict.get('bg_color', '00000000')
        )
    
    border = None
    if border_dict:
        def make_side(side_dict):
            if side_dict:
                return Side(style=side_dict)
            return Side()
        
        border = Border(
            left=make_side(border_dict.get('left')),
            right=make_side(border_dict.get('right')),
            top=make_side(border_dict.get('top')),
            bottom=make_side(border_dict.get('bottom'))
        )
    
    style_tuple = (font, alignment, fill, border, number_format or 'General')
    _style_cache[style_key] = style_tuple
    
    return style_tuple


def apply_cell_style(cell, cell_def):
    """Apply styling from a cell definition to an openpyxl cell."""
    font, alignment, fill, border, number_format = get_cached_style(
        cell_def.get('font'),
        cell_def.get('alignment'),
        cell_def.get('fill'),
        cell_def.get('border'),
        cell_def.get('number_format')
    )
    
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def create_workbook_from_json(value_overrides=None):
    """
    Create a new Excel workbook from template_structure.json (NO .xlsx template used).
    Builds workbook from scratch using JSON schema for complete control and lower memory.
    
    Args:
        value_overrides: Dict mapping cell coordinates (e.g., 'H2') to new values
                        for dynamic content like subscriber names, dates, etc.
    
    Returns:
        openpyxl.Workbook instance with structure/styling applied from JSON
    """
    template = load_template_structure()
    value_overrides = value_overrides or {}
    
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Apply column widths
    col_widths = template.get('column_widths', {})
    for col_str, width in col_widths.items():
        col_letter = get_column_letter(int(col_str))
        ws.column_dimensions[col_letter].width = width
    
    # Apply row heights
    row_heights = template.get('row_heights', {})
    for row_str, height in row_heights.items():
        ws.row_dimensions[int(row_str)].height = height
    
    # Apply cell styles and values
    cells = template.get('cells', [])
    for cell_def in cells:
        cell_coord = cell_def.get('cell')
        if not cell_coord:
            continue
        
        cell = ws[cell_coord]
        
        # Set value (use override if provided, otherwise use template value)
        if cell_coord in value_overrides:
            cell.value = value_overrides[cell_coord]
        else:
            cell.value = cell_def.get('value')
        
        # Apply styling
        apply_cell_style(cell, cell_def)
    
    # Apply merged cells
    merged_cells = template.get('merged_cells', [])
    for merge_range in merged_cells:
        try:
            ws.merge_cells(merge_range)
        except Exception as e:
            logger.warning(f"Failed to merge cells {merge_range}: {e}")
    
    # Add logo if it exists
    logo_path = os.path.join(settings.BASE_DIR.parent, 'media', 'Images', 'FirstCentralAPPROVEDLogo.png')
    if os.path.exists(logo_path):
        try:
            img = OpenpyxlImage(logo_path)
            # Position logo at top-left (B1 area) - adjust size to fit
            img.width = 180  # Adjust based on your template
            img.height = 60
            ws.add_image(img, 'B1')
        except Exception as e:
            logger.warning(f"Failed to add logo: {e}")
    
    return wb


def prepare_dynamic_values(subscriber_name, start_date_display, end_date_display):
    """
    Prepare value overrides for dynamic content in the template.
    
    Returns:
        Dict mapping cell coordinates to values
    """
    return {
        'H2': f"FirstCentral NIGERIA - BILLING DETAILS - {subscriber_name}",
        'B6': f"Report Generated for Records Between {start_date_display} and {end_date_display}",
    }


def write_billing_data(ws, summary_bills, subscriber_name, logger):
    """
    Write billing summary data to the worksheet with proper styling from template.
    
    Args:
        ws: Worksheet object
        summary_bills: Dict with product counts and rates
        subscriber_name: Name of the subscriber
        logger: Logger instance
    """
    from decimal import Decimal
    
    template = load_template_structure()
    
    # Product row mappings (same as current implementation)
    products_to_bill = {
        12: ('consumer_snap_check', 'Consumer Snap Check'),
        13: ('consumer_basic_trace', 'Consumer Basic Trace'),
        14: ('consumer_basic_credit', 'Consumer Basic Credit'),
        15: ('consumer_detailed_credit', 'Consumer Detailed Credit'),
        16: ('xscore_consumer_detailed_credit', 'X-Score Consumer Detailed Credit'),
        17: ('commercial_basic_trace', 'Commercial Basic Trace'),
        18: ('commercial_detailed_credit', 'Commercial Detailed Credit'),
        20: ('enquiry_report', 'Enquiry Report'),
        22: ('consumer_dud_cheque', 'Consumer Dud Cheque'),
        23: ('commercial_dud_cheque', 'Commercial Dud Cheque'),
        25: ('director_basic_report', 'Director Basic Report'),
        26: ('director_detailed_report', 'Director Detailed Report'),
    }
    
    # Get template cell definitions for styling
    template_cells = {(c['row'], c['col']): c for c in template.get('cells', [])}
    
    total_amount = Decimal('0.00')
    
    for row, (key, name) in products_to_bill.items():
        quantity = summary_bills.get(key, 0)
        rate = summary_bills.get(f'{key}_rate', Decimal('0.00'))
        amount = Decimal(quantity) * rate
        total_amount += amount
        
        # Write quantity to column I (9) with template styling
        cell_i = ws.cell(row=row, column=9)
        cell_i.value = quantity
        if (row, 9) in template_cells:
            apply_cell_style(cell_i, template_cells[(row, 9)])
        cell_i.number_format = '@'
        
        # Write rate to column M (13) with template styling
        cell_m = ws.cell(row=row, column=13)
        cell_m.value = f"₦{rate:,.2f}"
        if (row, 13) in template_cells:
            apply_cell_style(cell_m, template_cells[(row, 13)])
        cell_m.number_format = '@'
        
        # Write amount to column P (16) with template styling
        cell_p = ws.cell(row=row, column=16)
        cell_p.value = f"₦{amount:,.2f}"
        if (row, 16) in template_cells:
            apply_cell_style(cell_p, template_cells[(row, 16)])
        cell_p.number_format = '@'
    
    # Write totals with template styling
    vat_amount = total_amount * Decimal('0.075')
    amount_due = total_amount + vat_amount
    
    # Total amount (row 28, column P)
    cell_28 = ws.cell(row=28, column=16)
    cell_28.value = f"₦{total_amount:,.2f}"
    if (28, 16) in template_cells:
        apply_cell_style(cell_28, template_cells[(28, 16)])
    cell_28.number_format = '@'
    
    # VAT amount (row 29, column P)
    cell_29 = ws.cell(row=29, column=16)
    cell_29.value = f"₦{vat_amount:,.2f}"
    if (29, 16) in template_cells:
        apply_cell_style(cell_29, template_cells[(29, 16)])
    cell_29.number_format = '@'
    
    # Amount due (row 30, column P)
    cell_30 = ws.cell(row=30, column=16)
    cell_30.value = f"₦{amount_due:,.2f}"
    if (30, 16) in template_cells:
        apply_cell_style(cell_30, template_cells[(30, 16)])
    cell_30.number_format = '@'


def add_product_section_header(ws, row, product_name):
    """
    Add a product section header at the specified row.
    Replicates the styling and structure from rows 32-35 of the JSON template.
    
    Args:
        ws: Worksheet object
        row: Starting row number for the header
        product_name: Name of the product to display
    """
    template = load_template_structure()
    
    # Find cells in row 32-35 (header template)
    header_cells = [c for c in template.get('cells', []) if 32 <= c.get('row', 0) <= 35]
    
    # Calculate offset from template rows to target rows
    row_offset = row - 32
    
    # Apply header structure to new row
    for cell_def in header_cells:
        template_row = cell_def.get('row')
        col = cell_def.get('col')
        target_row = template_row + row_offset
        
        target_cell = ws.cell(row=target_row, column=col)
        
        # Override product name in the appropriate cell
        if 'product' in str(cell_def.get('value', '')).lower():
            target_cell.value = product_name
        else:
            target_cell.value = cell_def.get('value')
        
        apply_cell_style(target_cell, cell_def)
    
    # Apply merged cells for header section with proper offset
    merged_cells = template.get('merged_cells', [])
    for merge_range in merged_cells:
        # Parse merge range (e.g., "B32:H32" -> start_col=2, start_row=32, end_col=8, end_row=32)
        try:
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(merge_range)
            
            # Check if this merge is in the header rows (32-35)
            if 32 <= min_row <= 35:
                # Offset the merge range to the new position
                new_min_row = min_row + row_offset
                new_max_row = max_row + row_offset
                
                # Create new merge range at offset position
                new_merge_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
                
                try:
                    ws.merge_cells(new_merge_range)
                except Exception as e:
                    logger.warning(f"Failed to merge cells {new_merge_range}: {e}")
        except Exception as e:
            logger.warning(f"Failed to parse merge range {merge_range}: {e}")


def clone_template_row_to_data_row(ws, template_row_num, target_row_num, value_overrides=None):
    """
    Clone a template row's styling and structure to a target data row.
    Used for product detail rows that need consistent formatting.
    
    Args:
        ws: Worksheet object
        template_row_num: Row number in template to copy from (e.g., 36 for first data row)
        target_row_num: Target row number to copy to
        value_overrides: Dict mapping column numbers to values {col: value}
    """
    from openpyxl.utils import range_boundaries
    
    template = load_template_structure()
    value_overrides = value_overrides or {}
    
    # Get all cells from the template row
    template_cells = [c for c in template.get('cells', []) if c.get('row') == template_row_num]
    
    # If template row doesn't exist (like row 36), create basic data row styling
    if not template_cells:
        # Define standard data row styling based on typical template patterns
        data_row_style = {
            'font': {'name': 'Trebuchet MS', 'size': 8.0, 'bold': False, 'italic': False, 'color': 'FF000000'},
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'fill': {'pattern': 'solid', 'fg_color': 'FFFFFFFF', 'bg_color': '00000000'},
            'border': {
                'left': 'thin',
                'right': 'thin',
                'top': 'thin',
                'bottom': 'thin'
            },
            'number_format': 'General'
        }
        
        # Write values and apply styling to specified columns
        for col, value in value_overrides.items():
            target_cell = ws.cell(row=target_row_num, column=col)
            target_cell.value = value
            apply_cell_style(target_cell, data_row_style)
        
        # Set row height
        ws.row_dimensions[target_row_num].height = 25
        
        # Apply merged cells for specific columns (matching original template pattern)
        merge_patterns = [
            (5, 6),   # E:F - SubscriberName
            (7, 9),   # G:I - SystemUser
            (12, 14), # L:N - DetailsViewedDate
            (15, 17)  # O:Q - SearchOutput
        ]
        
        for start_col, end_col in merge_patterns:
            if start_col <= end_col:
                merge_range = f"{get_column_letter(start_col)}{target_row_num}:{get_column_letter(end_col)}{target_row_num}"
                try:
                    ws.merge_cells(merge_range)
                except Exception as e:
                    logger.warning(f"Failed to merge cells {merge_range}: {e}")
        
        return
    
    # Original logic for when template row exists
    # Copy each cell's style and value
    for cell_def in template_cells:
        col = cell_def.get('col')
        target_cell = ws.cell(row=target_row_num, column=col)
        
        # Set value ONLY if provided in overrides (don't copy template's empty values)
        if col in value_overrides:
            target_cell.value = value_overrides[col]
        # If no override and template has a value, only copy if it's not empty/None
        elif cell_def.get('value'):
            target_cell.value = cell_def.get('value')
        
        # Apply styling from template
        apply_cell_style(target_cell, cell_def)
    
    # Copy merged cells from template row to target row
    merged_cells = template.get('merged_cells', [])
    row_offset = target_row_num - template_row_num
    
    for merge_range in merged_cells:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(merge_range)
            
            # Check if this merge is in the template row
            if min_row == template_row_num:
                # Create corresponding merge in target row
                new_min_row = min_row + row_offset
                new_max_row = max_row + row_offset
                new_merge_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
                ws.merge_cells(new_merge_range)
        except Exception as e:
            logger.warning(f"Failed to merge cells in data row: {e}")
